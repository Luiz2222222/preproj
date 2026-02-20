from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.db.models import Q
from django.conf import settings
from django.http import StreamingHttpResponse
from django.core.mail import EmailMessage as DjangoEmailMessage
from django.contrib.auth.hashers import check_password
import os
import shutil
import zipfile
import io
import logging
from collections import defaultdict

from .models import TCC, SolicitacaoOrientacao, DocumentoTCC, EventoTimeline, BancaFase1, MembroBanca, AvaliacaoFase1, AvaliacaoFase2
from .serializers import (
    TCCSerializer,
    SolicitacaoOrientacaoSerializer,
    DocumentoTCCSerializer,
    EventoTimelineSerializer,
    CriarTCCSerializer,
    BancaFase1Serializer,
    AtualizarBancaFase1Serializer,
    AvaliacaoFase1Serializer,
    AvaliacaoFase1EscritaSerializer,
    AgendamentoDefesaSerializer
)
from notificacoes.services import criar_notificacao, criar_notificacao_em_massa
from notificacoes.constants import TipoNotificacao, PrioridadeNotificacao
from .permissions import (
    IsTCCOwnerOrRelated,
    IsSolicitacaoOwnerOrCoordenador,
    IsDocumentoTCCRelated,
    IsEventoTimelineVisible,
    IsAluno,
    IsCoordenador,
    IsProfessorOrCoordenador,
    IsProfessorCoordenadorOuAvaliador
)
from .constants import StatusSolicitacao, EtapaTCC, TipoEvento, Visibilidade, StatusDocumento, TipoDocumento
from .services import calcular_permissoes_tcc
from definicoes.models import CalendarioSemestre


def _pref(usuario, campo_prof):
    """Retorna o campo de preferência correto: prof_* ou aval_* conforme o tipo do usuário."""
    if usuario.tipo_usuario == 'AVALIADOR':
        return campo_prof.replace('prof_', 'aval_', 1)
    return campo_prof


class TCCViewSet(viewsets.ModelViewSet):
    """ViewSet para gerenciar TCCs."""

    queryset = TCC.objects.all()
    serializer_class = TCCSerializer
    permission_classes = [IsAuthenticated, IsTCCOwnerOrRelated]

    def get_queryset(self):
        """Filtrar TCCs conforme tipo de usuário."""
        usuario = self.request.user

        if usuario.tipo_usuario == 'COORDENADOR':
            # Coordenador vê todos
            return TCC.objects.all()

        elif usuario.tipo_usuario == 'ALUNO':
            # Aluno vê apenas seus TCCs
            return TCC.objects.filter(aluno=usuario)

        elif usuario.tipo_usuario == 'PROFESSOR':
            # Professor vê TCCs onde é orientador, coorientador OU avaliador na banca
            from .models import MembroBanca, BancaFase1

            # TCCs onde é orientador ou coorientador
            queryset = TCC.objects.filter(
                Q(orientador=usuario) | Q(coorientador=usuario)
            )

            # TCCs onde é membro de banca (avaliador)
            bancas_ids = MembroBanca.objects.filter(
                usuario=usuario,
                tipo='AVALIADOR'
            ).values_list('banca_id', flat=True)

            tccs_banca_ids = BancaFase1.objects.filter(
                id__in=bancas_ids
            ).values_list('tcc_id', flat=True)

            # Combinar os querysets
            queryset = queryset | TCC.objects.filter(id__in=tccs_banca_ids)

            return queryset.distinct()

        elif usuario.tipo_usuario == 'AVALIADOR':
            # Avaliador externo vê TCCs onde é coorientador OU membro de banca
            from .models import MembroBanca, BancaFase1

            # TCCs onde é coorientador
            queryset = TCC.objects.filter(coorientador=usuario)

            # TCCs onde é membro de banca
            bancas_ids = MembroBanca.objects.filter(
                usuario=usuario,
                tipo='AVALIADOR'
            ).values_list('banca_id', flat=True)

            tccs_banca_ids = BancaFase1.objects.filter(
                id__in=bancas_ids
            ).values_list('tcc_id', flat=True)

            # Combinar os querysets
            queryset = queryset | TCC.objects.filter(id__in=tccs_banca_ids)

            return queryset.distinct()

        return TCC.objects.none()

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def meu(self, request):
        """
        GET /api/tccs/meu/
        Retorna TCC(s) do usuário logado conforme seu tipo.
        """
        usuario = request.user

        if usuario.tipo_usuario == 'ALUNO':
            # Aluno: retorna seu TCC do semestre atual (ou mais recente)
            tcc = TCC.objects.filter(aluno=usuario).order_by('-criado_em').first()
            if tcc:
                serializer = self.get_serializer(tcc)
                return Response(serializer.data)

            # Se não tem TCC, verificar se há HistoricoRecusa
            from .models import HistoricoRecusa
            try:
                recusa = HistoricoRecusa.objects.get(aluno=usuario)
                return Response({
                    'status': 'RECUSADO',
                    'recusa': {
                        'parecer': recusa.parecer,
                        'coordenador_nome': recusa.coordenador_nome,
                        'recusado_em': recusa.recusado_em.isoformat()
                    }
                }, status=status.HTTP_200_OK)
            except HistoricoRecusa.DoesNotExist:
                return Response({'detail': 'Você ainda não possui TCC cadastrado'}, status=status.HTTP_404_NOT_FOUND)

        elif usuario.tipo_usuario == 'COORDENADOR':
            # Coordenador: retorna todos os TCCs
            tccs = TCC.objects.all().order_by('-criado_em')
            serializer = self.get_serializer(tccs, many=True)
            return Response(serializer.data)

        elif usuario.tipo_usuario == 'PROFESSOR':
            # Professor: retorna lista de orientandos (APENAS onde é orientador principal)
            # Co-orientações ficam em endpoint separado /tccs/minhas-coorientacoes/
            tccs = TCC.objects.filter(orientador=usuario).order_by('-criado_em')
            serializer = self.get_serializer(tccs, many=True)
            return Response(serializer.data)

        return Response({'detail': 'Tipo de usuário não autorizado'}, status=status.HTTP_403_FORBIDDEN)

    @action(detail=False, methods=['get'], permission_classes=[IsProfessorCoordenadorOuAvaliador])
    def avaliar(self, request):
        """
        GET /api/tccs/avaliar/
        Retorna TCCs que o usuário (professor, coordenador ou avaliador) precisa avaliar (onde é membro de banca).
        Inclui tanto Fase I (AVALIACAO_FASE_1) quanto Fase II (APRESENTACAO_FASE_2).
        Adiciona campos minha_avaliacao_fase1_status e minha_avaliacao_fase2_status.
        """
        usuario = request.user

        # Buscar TCCs onde o usuário é AVALIADOR na banca (não orientador)
        bancas_ids = MembroBanca.objects.filter(
            usuario=usuario,
            tipo='AVALIADOR'
        ).values_list('banca_id', flat=True)

        tccs_ids = BancaFase1.objects.filter(
            id__in=bancas_ids,
            status='COMPLETA'  # Apenas bancas já concluídas
        ).values_list('tcc_id', flat=True)

        # Retornar todos os TCCs onde o usuário participa de banca completa
        tccs = TCC.objects.filter(
            id__in=tccs_ids
        ).order_by('-criado_em')

        serializer = self.get_serializer(tccs, many=True)
        data = serializer.data

        # Adicionar status das avaliações do usuário atual
        for item in data:
            tcc_id = item['id']
            aval_f1 = AvaliacaoFase1.objects.filter(tcc_id=tcc_id, avaliador=usuario).first()
            aval_f2 = AvaliacaoFase2.objects.filter(tcc_id=tcc_id, avaliador=usuario).first()
            item['minha_avaliacao_fase1_status'] = aval_f1.status if aval_f1 else None
            item['minha_avaliacao_fase2_status'] = aval_f2.status if aval_f2 else None

        return Response(data)

    @action(detail=False, methods=['get'], url_path='minhas-coorientacoes', permission_classes=[IsProfessorCoordenadorOuAvaliador])
    def minhas_coorientacoes(self, request):
        """
        GET /api/tccs/minhas-coorientacoes/
        Retorna TCCs onde o usuário é co-orientador (não orientador principal).
        Para PROFESSOR, COORDENADOR e AVALIADOR.
        """
        usuario = request.user

        # Buscar TCCs onde é co-orientador (mas não orientador principal)
        tccs = TCC.objects.filter(coorientador=usuario).exclude(orientador=usuario).order_by('-criado_em')
        serializer = self.get_serializer(tccs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], permission_classes=[IsAluno], serializer_class=CriarTCCSerializer)
    def criar_com_solicitacao(self, request):
        """
        POST /api/tccs/criar_com_solicitacao/
        Cria TCC e envia solicitação de orientação (apenas aluno).
        """
        from django.db import transaction

        serializer = CriarTCCSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)

        dados = serializer.validated_data
        aluno = request.user
        professor = dados.pop('professor')

        # Validar que aluno não possui solicitação pendente em nenhum TCC
        solicitacao_pendente = SolicitacaoOrientacao.objects.filter(
            tcc__aluno=aluno,
            status=StatusSolicitacao.PENDENTE
        ).first()

        if solicitacao_pendente:
            return Response({
                'detail': 'Você já possui uma solicitação de orientação pendente. Aguarde a resposta do coordenador antes de criar outra.',
                'solicitacao_pendente_id': solicitacao_pendente.id,
                'tcc_id': solicitacao_pendente.tcc.id
            }, status=status.HTTP_400_BAD_REQUEST)

        # Validar prazo de envio de documentos
        semestre = dados.get('semestre')
        calendario = CalendarioSemestre.obter_calendario_atual(semestre)
        if calendario and calendario.envio_documentos_fim:
            hoje = timezone.localdate()
            if hoje > calendario.envio_documentos_fim:
                return Response({
                    'detail': 'Período de envio de documentos encerrado. Solicite ao coordenador a liberação para casos excepcionais.'
                }, status=status.HTTP_400_BAD_REQUEST)

        # Usar transação atômica para garantir consistência
        with transaction.atomic():
            # Limpar HistoricoRecusa do aluno (se existir)
            from .models import HistoricoRecusa
            HistoricoRecusa.objects.filter(aluno=aluno).delete()

            # Criar TCC (sem orientador e sem coorientador FK — serão setados na aprovação)
            tcc = TCC.objects.create(
                aluno=aluno,
                titulo=dados['titulo'],
                resumo=dados.get('resumo', ''),
                semestre=dados['semestre'],
                coorientador_nome=dados.get('coorientador_nome', ''),
                coorientador_titulacao=dados.get('coorientador_titulacao', ''),
                coorientador_afiliacao=dados.get('coorientador_afiliacao', ''),
                coorientador_lattes=dados.get('coorientador_lattes', '')
            )

            # Criar evento de criação
            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=aluno,
                tipo_evento=TipoEvento.CRIACAO_TCC,
                descricao=f'TCC "{tcc.titulo}" criado por {aluno.nome_completo}',
                visibilidade=Visibilidade.TODOS
            )

            # Preparar dados do co-orientador para a solicitação
            coorientador_interno = dados.get('coorientador')
            if coorientador_interno:
                # Co-orientador interno: pegar dados do usuário
                coorientador_nome = coorientador_interno.nome_completo
                # Se tratamento for "Outro", usar o valor customizado
                if coorientador_interno.tratamento == 'Outro' and coorientador_interno.tratamento_customizado:
                    coorientador_titulacao = coorientador_interno.tratamento_customizado
                else:
                    coorientador_titulacao = coorientador_interno.tratamento or ''
                # Se afiliacao for "Outro", usar o valor customizado
                if coorientador_interno.afiliacao == 'Outro' and coorientador_interno.afiliacao_customizada:
                    coorientador_afiliacao = coorientador_interno.afiliacao_customizada
                else:
                    coorientador_afiliacao = coorientador_interno.departamento or coorientador_interno.afiliacao or ''
                coorientador_lattes = ''
            else:
                # Co-orientador externo: usar dados informados
                coorientador_nome = dados.get('coorientador_nome', '')
                coorientador_titulacao = dados.get('coorientador_titulacao', '')
                coorientador_afiliacao = dados.get('coorientador_afiliacao', '')
                coorientador_lattes = dados.get('coorientador_lattes', '')

            # Criar solicitação de orientação
            solicitacao = SolicitacaoOrientacao.objects.create(
                tcc=tcc,
                professor=professor,
                mensagem=dados.get('mensagem', ''),
                coorientador_nome=coorientador_nome,
                coorientador_titulacao=coorientador_titulacao,
                coorientador_afiliacao=coorientador_afiliacao,
                coorientador_lattes=coorientador_lattes
            )

            # Criar evento de solicitação
            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=aluno,
                tipo_evento=TipoEvento.SOLICITACAO_ENVIADA,
                descricao=f'Solicitação de orientação enviada para {professor.nome_completo}',
                detalhes_json={'professor_id': professor.id, 'solicitacao_id': solicitacao.id},
                visibilidade=Visibilidade.TODOS
            )

        # Documentos obrigatórios conforme Esquema.txt
        documentos_obrigatorios = [
            {
                'tipo': 'PLANO_DESENVOLVIMENTO',
                'tipo_display': 'Plano de Desenvolvimento',
                'obrigatorio': True,
                'descricao': 'Documento obrigatório para início do TCC'
            },
            {
                'tipo': 'TERMO_ACEITE',
                'tipo_display': 'Termo de Aceite de Orientação',
                'obrigatorio': True,
                'descricao': 'Termo assinado pelo orientador'
            }
        ]

        # Verificar pendências (documentos já enviados)
        documentos_enviados = DocumentoTCC.objects.filter(tcc=tcc).values_list('tipo_documento', flat=True)
        pendencias = []

        for doc in documentos_obrigatorios:
            if doc['tipo'] not in documentos_enviados:
                pendencias.append({
                    'tipo': doc['tipo'],
                    'tipo_display': doc['tipo_display'],
                    'mensagem': f"{doc['tipo_display']} ainda não foi enviado"
                })

        return Response({
            'tcc': TCCSerializer(tcc, context={'request': request}).data,
            'solicitacao': SolicitacaoOrientacaoSerializer(solicitacao, context={'request': request}).data,
            'documentos_obrigatorios': documentos_obrigatorios,
            'pendencias': pendencias,
            'message': 'TCC criado e solicitação enviada com sucesso'
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], permission_classes=[IsAluno])
    def solicitacoes(self, request, pk=None):
        """
        POST /api/tccs/{id}/solicitacoes/
        Cria solicitação de orientação para este TCC (rota aninhada).
        """
        tcc = self.get_object()

        # Validar que é o aluno dono
        if tcc.aluno != request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Você não pode criar solicitação para TCC de outro aluno')

        # Validar que não existe solicitação pendente
        if SolicitacaoOrientacao.objects.filter(tcc=tcc, status=StatusSolicitacao.PENDENTE).exists():
            return Response(
                {'detail': 'Já existe uma solicitação pendente para este TCC'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obter professor do payload
        professor_id = request.data.get('professor')
        if not professor_id:
            return Response(
                {'detail': 'Campo professor é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from users.models import Usuario
        try:
            professor = Usuario.objects.get(id=professor_id)
            if professor.tipo_usuario not in ['PROFESSOR', 'COORDENADOR']:
                return Response(
                    {'detail': 'Professor deve ser do tipo PROFESSOR ou COORDENADOR'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Usuario.DoesNotExist:
            return Response(
                {'detail': 'Professor não encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Criar solicitação
        solicitacao = SolicitacaoOrientacao.objects.create(
            tcc=tcc,
            professor=professor,
            mensagem=request.data.get('mensagem', ''),
            coorientador_nome=request.data.get('coorientador_nome', ''),
            coorientador_titulacao=request.data.get('coorientador_titulacao', ''),
            coorientador_afiliacao=request.data.get('coorientador_afiliacao', ''),
            coorientador_lattes=request.data.get('coorientador_lattes', '')
        )

        # Criar evento
        EventoTimeline.objects.create(
            tcc=tcc,
            usuario=request.user,
            tipo_evento=TipoEvento.SOLICITACAO_ENVIADA,
            descricao=f'Solicitação de orientação enviada para {professor.nome_completo}',
            detalhes_json={'professor_id': professor.id, 'solicitacao_id': solicitacao.id},
            visibilidade=Visibilidade.TODOS
        )

        return Response(
            SolicitacaoOrientacaoSerializer(solicitacao, context={'request': request}).data,
            status=status.HTTP_201_CREATED
        )

    @action(detail=True, methods=['get', 'post'], permission_classes=[IsAuthenticated])
    def documentos(self, request, pk=None):
        """
        GET/POST /api/tccs/{id}/documentos/
        Lista ou envia documentos para este TCC (rota aninhada).
        """
        tcc = self.get_object()

        if request.method == 'GET':
            # Listar documentos do TCC
            usuario = request.user

            # Verificar se usuário é avaliador da banca
            eh_avaliador = False
            if usuario.tipo_usuario in ['PROFESSOR', 'COORDENADOR']:
                from .models import MembroBanca, BancaFase1
                # Verificar se é membro avaliador da banca deste TCC
                eh_avaliador = MembroBanca.objects.filter(
                    banca__tcc=tcc,
                    usuario=usuario,
                    tipo='AVALIADOR'
                ).exists()

            # Se é avaliador, verificar se existe documento anônimo
            if eh_avaliador:
                try:
                    banca = BancaFase1.objects.get(tcc=tcc)
                    if banca.documento_avaliacao:
                        # Retornar APENAS o documento anônimo para avaliação duplo-cega
                        documentos = DocumentoTCC.objects.filter(id=banca.documento_avaliacao.id)
                    else:
                        # Se não há documento anônimo, retornar monografia original
                        documentos = DocumentoTCC.objects.filter(
                            tcc=tcc,
                            tipo_documento=TipoDocumento.MONOGRAFIA
                        ).order_by('-criado_em')[:1]  # Apenas a mais recente
                except BancaFase1.DoesNotExist:
                    # Se não tem banca, retornar vazio para avaliador
                    documentos = DocumentoTCC.objects.none()
            else:
                # Para outros usuários (aluno, orientador, coordenador), retornar todos
                documentos = DocumentoTCC.objects.filter(tcc=tcc).order_by('-criado_em')

            serializer = DocumentoTCCSerializer(documentos, many=True, context={'request': request})
            return Response(serializer.data)

        elif request.method == 'POST':
            # Upload de documento
            usuario = request.user

            # Validar permissão
            tem_permissao = False
            if usuario.tipo_usuario == 'COORDENADOR':
                tem_permissao = True
            elif usuario.tipo_usuario == 'ALUNO' and tcc.aluno == usuario:
                tem_permissao = True
            elif usuario.tipo_usuario in ['PROFESSOR', 'COORDENADOR']:
                if tcc.orientador == usuario or tcc.coorientador == usuario:
                    tem_permissao = True

            if not tem_permissao:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied('Você não tem permissão para enviar documentos neste TCC')

            # Criar documento (tcc preenchido automaticamente)
            # Copiar dados e garantir que tipo_documento seja string (não lista)
            dados = request.data.copy()
            tipo_documento = request.data.get('tipo_documento')
            if isinstance(tipo_documento, list):
                tipo_documento = tipo_documento[0]
            dados['tipo_documento'] = tipo_documento
            dados['tcc'] = tcc.id

            serializer = DocumentoTCCSerializer(data=dados, context={'request': request})
            serializer.is_valid(raise_exception=True)

            # Validar prazos baseados no tipo de documento (coordenador bypassa)
            if usuario.tipo_usuario != 'COORDENADOR':
                permissoes = calcular_permissoes_tcc(tcc)

                # Documentos iniciais
                if tipo_documento in [TipoDocumento.PLANO_DESENVOLVIMENTO, TipoDocumento.TERMO_ACEITE]:
                    if not permissoes['pode_enviar_documentos_iniciais']:
                        from rest_framework.exceptions import PermissionDenied
                        raise PermissionDenied('Período de envio de documentos iniciais encerrado. Solicite liberação ao coordenador.')

                # Monografia
                elif tipo_documento == TipoDocumento.MONOGRAFIA:
                    if not permissoes['pode_enviar_monografia']:
                        from rest_framework.exceptions import PermissionDenied
                        raise PermissionDenied('Período de envio de monografia encerrado. Solicite liberação ao coordenador.')

                # Termo de Solicitação de Avaliação
                elif tipo_documento == TipoDocumento.TERMO_SOLICITACAO_AVALIACAO:
                    if not permissoes['pode_solicitar_avaliacao']:
                        from rest_framework.exceptions import PermissionDenied
                        raise PermissionDenied('Período de solicitação de avaliação encerrado. Solicite liberação ao coordenador.')

            documento = serializer.save(enviado_por=usuario, tcc=tcc)

            # Criar evento
            # Nota: .capitalize() padroniza para sentence case (apenas primeira letra maiúscula).
            # Eventos antigos continuam com formato original - apenas novos seguem este padrão.
            tipo_display = documento.get_tipo_documento_display().capitalize()
            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=usuario,
                tipo_evento=TipoEvento.UPLOAD_DOCUMENTO,
                descricao=f'Documento enviado: {tipo_display} (v{documento.versao})',
                detalhes_json={
                    'documento_id': documento.id,
                    'tipo': documento.tipo_documento,
                    'versao': documento.versao
                },
                visibilidade=Visibilidade.TODOS
            )

            return Response(DocumentoTCCSerializer(documento, context={'request': request}).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], permission_classes=[IsProfessorOrCoordenador])
    def confirmar_continuidade(self, request, pk=None):
        """
        POST /api/tccs/{id}/confirmar_continuidade/
        Professor/Coorientador confirma continuidade do aluno no TCC.
        Permitido se: liberação manual, monografia aprovada (antecipa), ou dentro do prazo.
        """
        tcc = self.get_object()
        usuario = request.user

        # Validar que usuario é orientador ou coorientador
        if tcc.orientador != usuario and tcc.coorientador != usuario:
            return Response(
                {'detail': 'Apenas orientador ou coorientador pode confirmar continuidade'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Validar que TCC está em DESENVOLVIMENTO
        if tcc.etapa_atual != EtapaTCC.DESENVOLVIMENTO:
            return Response(
                {'detail': f'TCC deve estar na etapa DESENVOLVIMENTO. Etapa atual: {tcc.get_etapa_atual_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que continuidade ainda não foi confirmada
        if tcc.flag_continuidade:
            return Response(
                {'detail': 'Continuidade já foi confirmada anteriormente'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar prazo de confirmação de continuidade
        permissoes = calcular_permissoes_tcc(tcc)
        if not permissoes['pode_confirmar_continuidade']:
            return Response(
                {'detail': 'Não é possível confirmar continuidade fora do prazo. Solicite liberação ao coordenador.'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Confirmar continuidade
        tcc.flag_continuidade = True
        tcc.save()

        # Criar evento
        EventoTimeline.objects.create(
            tcc=tcc,
            usuario=usuario,
            tipo_evento=TipoEvento.APROVACAO_CONTINUIDADE,
            descricao=f'Continuidade aprovada por {usuario.nome_completo}. Aluno pode prosseguir para avaliação',
            detalhes_json={'aprovador_id': usuario.id},
            visibilidade=Visibilidade.TODOS
        )

        # Criar notificações COM e-mail
        from notificacoes.services import criar_notificacao_com_email, criar_notificacao_em_massa_com_email

        # Notificar aluno
        criar_notificacao_com_email(
            usuario=tcc.aluno,
            tipo=TipoNotificacao.CONTINUIDADE_APROVADA,
            titulo='Continuidade Aprovada',
            mensagem=f'Sua continuidade foi aprovada por {usuario.nome_completo}. Agora você pode prosseguir para a avaliação.',
            campo_preferencia='aluno_continuidade_aprovada',
            action_url=f'/tccs/{tcc.id}',
            tcc_id=tcc.id,
            prioridade=PrioridadeNotificacao.ALTA
        )

        # Notificar professor (orientador/coorientador)
        criar_notificacao_com_email(
            usuario=usuario,
            tipo=TipoNotificacao.CONTINUIDADE_APROVADA,
            titulo='Continuidade Confirmada',
            mensagem=f'Você aprovou a continuidade do TCC "{tcc.titulo}".',
            campo_preferencia='prof_continuidade_aprovada',
            action_url=f'/tccs/{tcc.id}',
            tcc_id=tcc.id,
            prioridade=PrioridadeNotificacao.NORMAL
        )

        # Notificar coordenador
        from users.models import Usuario
        coordenadores = Usuario.objects.filter(tipo_usuario='COORDENADOR')
        criar_notificacao_em_massa_com_email(
            usuarios=list(coordenadores),
            tipo=TipoNotificacao.CONTINUIDADE_APROVADA,
            titulo='Continuidade Aprovada',
            mensagem=f'Continuidade do TCC "{tcc.titulo}" foi aprovada por {usuario.nome_completo}.',
            campo_preferencia='coord_continuidade_aprovada',
            action_url=f'/tccs/{tcc.id}',
            tcc_id=tcc.id,
            prioridade=PrioridadeNotificacao.NORMAL
        )

        return Response({
            'message': 'Continuidade confirmada com sucesso',
            'flag_continuidade': tcc.flag_continuidade,
            'etapa_atual': tcc.etapa_atual
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], permission_classes=[IsProfessorOrCoordenador])
    def rejeitar_continuidade(self, request, pk=None):
        """
        POST /api/tccs/{id}/rejeitar_continuidade/
        Professor/Coorientador rejeita continuidade do aluno no TCC.
        Move o TCC para DESCONTINUADO.
        """
        tcc = self.get_object()
        usuario = request.user

        # Validar que usuario é orientador ou coorientador
        if tcc.orientador != usuario and tcc.coorientador != usuario:
            return Response(
                {'detail': 'Apenas orientador ou coorientador pode rejeitar continuidade'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Validar que TCC está em DESENVOLVIMENTO
        if tcc.etapa_atual != EtapaTCC.DESENVOLVIMENTO:
            return Response(
                {'detail': f'TCC deve estar na etapa DESENVOLVIMENTO. Etapa atual: {tcc.get_etapa_atual_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que continuidade ainda não foi confirmada
        if tcc.flag_continuidade:
            return Response(
                {'detail': 'Continuidade já foi confirmada. Não é possível rejeitar.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar prazo de confirmação de continuidade
        permissoes = calcular_permissoes_tcc(tcc)
        if not permissoes['pode_confirmar_continuidade']:
            return Response(
                {'detail': 'Não é possível rejeitar continuidade fora do prazo. Solicite liberação ao coordenador.'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Rejeitar continuidade - mover para DESCONTINUADO
        tcc.flag_continuidade = False
        tcc.etapa_atual = EtapaTCC.DESCONTINUADO
        tcc.save()

        # Criar evento
        EventoTimeline.objects.create(
            tcc=tcc,
            usuario=usuario,
            tipo_evento=TipoEvento.REPROVACAO_CONTINUIDADE,
            descricao=f'Continuidade rejeitada por {usuario.nome_completo}. TCC descontinuado.',
            detalhes_json={'rejeitador_id': usuario.id},
            visibilidade=Visibilidade.TODOS
        )

        # Criar notificações COM e-mail
        from notificacoes.services import criar_notificacao_com_email, criar_notificacao_em_massa_com_email

        # Notificar aluno
        criar_notificacao_com_email(
            usuario=tcc.aluno,
            tipo=TipoNotificacao.CONTINUIDADE_REJEITADA,
            titulo='Continuidade Rejeitada',
            mensagem=f'Sua continuidade foi rejeitada por {usuario.nome_completo}. O TCC foi descontinuado.',
            campo_preferencia='aluno_continuidade_rejeitada',
            action_url=f'/tccs/{tcc.id}',
            tcc_id=tcc.id,
            prioridade=PrioridadeNotificacao.ALTA
        )

        # Notificar coordenador
        from users.models import Usuario
        coordenadores = Usuario.objects.filter(tipo_usuario='COORDENADOR')
        criar_notificacao_em_massa_com_email(
            usuarios=list(coordenadores),
            tipo=TipoNotificacao.CONTINUIDADE_REJEITADA,
            titulo='Continuidade Rejeitada',
            mensagem=f'Continuidade do TCC "{tcc.titulo}" foi rejeitada por {usuario.nome_completo}. TCC descontinuado.',
            campo_preferencia='coord_continuidade_rejeitada',
            action_url=f'/tccs/{tcc.id}',
            tcc_id=tcc.id,
            prioridade=PrioridadeNotificacao.NORMAL
        )

        return Response({
            'message': 'Continuidade rejeitada. TCC descontinuado.',
            'flag_continuidade': tcc.flag_continuidade,
            'etapa_atual': tcc.etapa_atual
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], permission_classes=[IsProfessorOrCoordenador])
    def enviar_termo_avaliacao(self, request, pk=None):
        """
        POST /api/tccs/{id}/enviar-termo-avaliacao/
        Professor/Coorientador envia Termo de Solicitação de Avaliação.
        Valida flag_continuidade=True, cria documento, seta flag_liberado_avaliacao=True
        e move TCC para FORMACAO_BANCA_FASE_1.
        """
        from django.db import transaction

        tcc = self.get_object()
        usuario = request.user

        # Validar que usuario é orientador ou coorientador
        if tcc.orientador != usuario and tcc.coorientador != usuario:
            return Response(
                {'detail': 'Apenas orientador ou coorientador pode enviar termo de avaliação'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Validar que TCC está em DESENVOLVIMENTO
        if tcc.etapa_atual != EtapaTCC.DESENVOLVIMENTO:
            return Response(
                {'detail': f'TCC deve estar na etapa DESENVOLVIMENTO. Etapa atual: {tcc.get_etapa_atual_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar flag_continuidade=True
        if not tcc.flag_continuidade:
            return Response(
                {'detail': 'Continuidade deve ser confirmada antes de enviar termo de avaliação'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar monografia aprovada
        monografia_recente = DocumentoTCC.objects.filter(
            tcc=tcc,
            tipo_documento=TipoDocumento.MONOGRAFIA
        ).order_by('-criado_em').first()

        if not monografia_recente or monografia_recente.status != StatusDocumento.APROVADO:
            return Response(
                {'detail': 'Monografia mais recente deve estar APROVADA'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar prazo de solicitação de avaliação
        permissoes = calcular_permissoes_tcc(tcc)
        if not permissoes['pode_solicitar_avaliacao']:
            return Response(
                {'detail': 'Não é possível solicitar avaliação fora do prazo. Solicite liberação ao coordenador.'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Validar que arquivo foi enviado
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            return Response(
                {'detail': 'Campo arquivo é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Usar transação atômica
        with transaction.atomic():
            # Criar documento
            documento = DocumentoTCC.objects.create(
                tcc=tcc,
                tipo_documento=TipoDocumento.TERMO_SOLICITACAO_AVALIACAO,
                arquivo=arquivo,
                nome_original=arquivo.name,
                tamanho=arquivo.size,
                enviado_por=usuario,
                status=StatusDocumento.PENDENTE
            )

            # Atualizar flags e etapa
            tcc.flag_liberado_avaliacao = True
            tcc.etapa_atual = EtapaTCC.FORMACAO_BANCA_FASE_1
            tcc.save()

            # Criar evento de liberação
            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=usuario,
                tipo_evento=TipoEvento.LIBERACAO_AVALIACAO,
                descricao=f'TCC liberado para avaliação por {usuario.nome_completo}. Termo de Solicitação enviado. Etapa: Formação de Banca - Fase 1',
                detalhes_json={
                    'liberador_id': usuario.id,
                    'documento_id': documento.id
                },
                visibilidade=Visibilidade.TODOS
            )

        return Response({
            'message': 'Termo de avaliação enviado com sucesso',
            'documento': DocumentoTCCSerializer(documento, context={'request': request}).data,
            'flag_liberado_avaliacao': tcc.flag_liberado_avaliacao,
            'etapa_atual': tcc.etapa_atual,
            'etapa_display': tcc.get_etapa_atual_display()
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get', 'put'], url_path='banca-fase1', permission_classes=[IsAuthenticated])
    def banca_fase1(self, request, pk=None):
        """
        GET/PUT /api/tccs/{id}/banca-fase1/
        GET: Retorna composição da banca (lazy create se não existir)
        PUT: Coordenador atualiza composição (avaliadores)
        """
        from django.db import transaction
        from users.models import Usuario

        tcc = self.get_object()

        if request.method == 'GET':
            # Lazy create: criar banca se não existir
            banca, created = BancaFase1.objects.get_or_create(tcc=tcc)

            # Garantir que orientador está na banca (evita duplicatas)
            if tcc.orientador:
                MembroBanca.objects.get_or_create(
                    banca=banca,
                    usuario=tcc.orientador,
                    tipo='ORIENTADOR',
                    defaults={
                        'indicado_por': 'ORIENTADOR',
                        'ordem': 0
                    }
                )

            serializer = BancaFase1Serializer(banca, context={'request': request})
            return Response(serializer.data)

        elif request.method == 'PUT':
            # Apenas coordenador pode atualizar
            if request.user.tipo_usuario != 'COORDENADOR':
                return Response(
                    {'detail': 'Apenas coordenador pode atualizar a composição da banca'},
                    status=status.HTTP_403_FORBIDDEN
                )

            # TCC deve estar em FORMACAO_BANCA_FASE_1
            if tcc.etapa_atual != EtapaTCC.FORMACAO_BANCA_FASE_1:
                return Response(
                    {'detail': f'TCC deve estar em FORMACAO_BANCA_FASE_1. Etapa atual: {tcc.get_etapa_atual_display()}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validar dados
            serializer = AtualizarBancaFase1Serializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)

            banca, _ = BancaFase1.objects.get_or_create(tcc=tcc)

            # Validar que orientador/coorientador não estão na lista de avaliadores
            novos_avaliadores_ids = set(serializer.validated_data['avaliadores'])
            membros_fixos = []
            if tcc.orientador and tcc.orientador.id in novos_avaliadores_ids:
                membros_fixos.append(f'{tcc.orientador.nome_completo} (orientador)')
            if tcc.coorientador and tcc.coorientador.id in novos_avaliadores_ids:
                membros_fixos.append(f'{tcc.coorientador.nome_completo} (coorientador)')

            if membros_fixos:
                return Response({
                    'detail': f'Orientador e coorientador não podem ser avaliadores. Membros inválidos: {", ".join(membros_fixos)}'
                }, status=status.HTTP_400_BAD_REQUEST)

            with transaction.atomic():
                # Limpar TODOS os membros existentes (avaliadores e orientador) para reconstruir
                MembroBanca.objects.filter(banca=banca).delete()

                # Recriar orientador
                if tcc.orientador:
                    MembroBanca.objects.create(
                        banca=banca,
                        usuario=tcc.orientador,
                        tipo='ORIENTADOR',
                        indicado_por='ORIENTADOR',
                        ordem=0
                    )

                # Adicionar novos avaliadores
                for idx, avaliador_id in enumerate(serializer.validated_data['avaliadores'], start=1):
                    avaliador = Usuario.objects.get(id=avaliador_id)
                    MembroBanca.objects.create(
                        banca=banca,
                        usuario=avaliador,
                        tipo='AVALIADOR',
                        indicado_por='COORDENADOR',
                        ordem=idx
                    )

            # Retornar banca atualizada
            serializer_response = BancaFase1Serializer(banca, context={'request': request})
            return Response(serializer_response.data)

    @action(detail=True, methods=['post'], url_path='banca-fase1/concluir', permission_classes=[IsCoordenador])
    def concluir_banca_fase1(self, request, pk=None):
        """
        POST /api/tccs/{id}/banca-fase1/concluir/
        Coordenador finaliza formação da banca:
        - Marca status COMPLETA
        - Cria avaliações PENDENTE para cada avaliador
        - Move TCC para AVALIACAO_FASE_1
        - Cria evento
        """
        from django.db import transaction

        tcc = self.get_object()

        # Validar etapa
        if tcc.etapa_atual != EtapaTCC.FORMACAO_BANCA_FASE_1:
            return Response(
                {'detail': f'TCC deve estar na etapa FORMACAO_BANCA_FASE_1. Etapa atual: {tcc.get_etapa_atual_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obter banca
        try:
            banca = BancaFase1.objects.get(tcc=tcc)
        except BancaFase1.DoesNotExist:
            return Response(
                {'detail': 'Banca ainda não foi criada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que tem pelo menos 2 avaliadores
        avaliadores = MembroBanca.objects.filter(banca=banca, tipo='AVALIADOR')
        if avaliadores.count() < 2:
            return Response(
                {'detail': 'Banca deve ter pelo menos 2 avaliadores'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que não está completa ainda
        if banca.status == 'COMPLETA':
            return Response(
                {'detail': 'Banca já foi concluída anteriormente'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Processar documento de avaliação
        arquivo_avaliacao = request.FILES.get('documento_avaliacao')

        # Usar transação atômica
        with transaction.atomic():
            documento_avaliacao = None

            if arquivo_avaliacao:
                # Coordenador enviou arquivo anônimo manualmente
                from .serializers import AtualizarBancaFase1Serializer
                serializer = AtualizarBancaFase1Serializer(data={'documento_avaliacao': arquivo_avaliacao}, partial=True)
                if serializer.is_valid():
                    documento_avaliacao = DocumentoTCC.objects.create(
                        tcc=tcc,
                        tipo_documento=TipoDocumento.MONOGRAFIA_AVALIACAO,
                        arquivo=arquivo_avaliacao,
                        nome_original=arquivo_avaliacao.name,
                        tamanho=arquivo_avaliacao.size,
                        enviado_por=request.user,
                        status=StatusDocumento.APROVADO
                    )
                else:
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            else:
                # Gerar versão sem as 2 primeiras páginas a partir da monografia do aluno
                monografia = DocumentoTCC.objects.filter(
                    tcc=tcc,
                    tipo_documento=TipoDocumento.MONOGRAFIA
                ).order_by('-versao').first()

                if not monografia or not monografia.arquivo:
                    return Response(
                        {'detail': 'Nenhuma monografia encontrada para este TCC. Envie uma versão anônima manualmente.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                try:
                    from django.core.files.base import ContentFile
                    from PyPDF2 import PdfReader, PdfWriter
                    import io
                    import tempfile
                    import os

                    nome_arquivo = (monografia.nome_original or '').lower()
                    eh_word = nome_arquivo.endswith('.doc') or nome_arquivo.endswith('.docx')

                    monografia.arquivo.open('rb')
                    conteudo_original = monografia.arquivo.read()
                    monografia.arquivo.close()

                    if eh_word:
                        # Word: converter para PDF primeiro usando docx2pdf
                        from docx2pdf import convert
                        import pythoncom

                        # Salvar Word em arquivo temporário
                        sufixo = '.docx' if nome_arquivo.endswith('.docx') else '.doc'
                        with tempfile.NamedTemporaryFile(suffix=sufixo, delete=False) as tmp_word:
                            tmp_word.write(conteudo_original)
                            tmp_word_path = tmp_word.name

                        tmp_pdf_path = tmp_word_path.rsplit('.', 1)[0] + '.pdf'

                        try:
                            pythoncom.CoInitialize()
                            convert(tmp_word_path, tmp_pdf_path)
                            with open(tmp_pdf_path, 'rb') as f:
                                pdf_bytes = f.read()
                        finally:
                            pythoncom.CoUninitialize()
                            # Limpar arquivos temporários
                            if os.path.exists(tmp_word_path):
                                os.unlink(tmp_word_path)
                            if os.path.exists(tmp_pdf_path):
                                os.unlink(tmp_pdf_path)

                        pdf_stream = io.BytesIO(pdf_bytes)
                    else:
                        # Já é PDF
                        pdf_stream = io.BytesIO(conteudo_original)

                    # Remover as 2 primeiras páginas do PDF
                    reader = PdfReader(pdf_stream)
                    writer = PdfWriter()

                    if len(reader.pages) <= 2:
                        return Response(
                            {'detail': 'A monografia tem 2 páginas ou menos.'},
                            status=status.HTTP_400_BAD_REQUEST
                        )

                    for i in range(2, len(reader.pages)):
                        writer.add_page(reader.pages[i])

                    buffer = io.BytesIO()
                    writer.write(buffer)
                    tamanho = buffer.tell()
                    buffer.seek(0)

                    nome_anonimo = f"avaliacao_{tcc.id}.pdf"
                    conteudo_file = ContentFile(buffer.read(), name=nome_anonimo)

                    documento_avaliacao = DocumentoTCC.objects.create(
                        tcc=tcc,
                        tipo_documento=TipoDocumento.MONOGRAFIA_AVALIACAO,
                        arquivo=conteudo_file,
                        nome_original=nome_anonimo,
                        tamanho=tamanho,
                        enviado_por=request.user,
                        status=StatusDocumento.APROVADO
                    )
                except Exception as e:
                    return Response(
                        {'detail': f'Erro ao processar monografia: {str(e)}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Atualizar banca
            banca.status = 'COMPLETA'
            banca.data_formacao = timezone.now()
            banca.formada_por = request.user
            banca.documento_avaliacao = documento_avaliacao
            banca.save()

            # Criar avaliações para cada avaliador
            for membro in avaliadores:
                AvaliacaoFase1.objects.get_or_create(
                    tcc=tcc,
                    avaliador=membro.usuario,
                    defaults={'status': 'PENDENTE'}
                )

            # Atualizar TCC
            tcc.etapa_atual = EtapaTCC.AVALIACAO_FASE_1
            tcc.save()

            # Criar evento
            nomes_avaliadores = ', '.join([m.usuario.nome_completo for m in avaliadores])
            descricao = f'Banca da Fase I formada por {request.user.nome_completo}. Avaliadores: {nomes_avaliadores}.'
            if documento_avaliacao:
                descricao += ' Documento anônimo para avaliação enviado.'
            descricao += ' Etapa: Avaliação - Fase 1'

            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=request.user,
                tipo_evento=TipoEvento.FORMACAO_BANCA,
                descricao=descricao,
                detalhes_json={
                    'banca_id': banca.id,
                    'avaliadores': [m.usuario.id for m in avaliadores],
                    'documento_avaliacao_id': documento_avaliacao.id if documento_avaliacao else None
                },
                visibilidade=Visibilidade.TODOS
            )

            # Notificar avaliadores por email
            from notificacoes.services import criar_notificacao_com_email
            from notificacoes.constants import TipoNotificacao, PrioridadeNotificacao

            for membro in avaliadores:
                prof = membro.usuario
                # Determinar campo de preferência conforme tipo de usuário
                campo_pref = 'aval_participacao_banca' if prof.tipo_usuario == 'AVALIADOR' else 'prof_participacao_banca'

                corpo_html = f"""
<p>Olá, {prof.nome_completo},</p>

<p>Você foi adicionado como avaliador em uma banca de avaliação de TCC.</p>

<p>A monografia para avaliação já está disponível no sistema. Acesse o sistema: <a href="http://localhost:5500">http://localhost:5500</a></p>

<p>---<br>
Portal TCC<br>
Esta é uma notificação automática. Para mais informações, acesse o sistema.</p>
"""
                corpo_texto = f"""Olá, {prof.nome_completo},

Você foi adicionado como avaliador em uma banca de avaliação de TCC.

A monografia para avaliação já está disponível no sistema. Acesse o sistema: http://localhost:5500

---
Portal TCC
Esta é uma notificação automática. Para mais informações, acesse o sistema.
"""
                criar_notificacao_com_email(
                    usuario=prof,
                    tipo=TipoNotificacao.DOCUMENTO_ENVIADO,
                    titulo='Participação em Banca de Avaliação',
                    mensagem='Você foi adicionado como avaliador em uma banca de avaliação de TCC.',
                    campo_preferencia=campo_pref,
                    action_url='/meus-tccs',
                    prioridade=PrioridadeNotificacao.ALTA,
                    corpo_html_customizado=corpo_html,
                    corpo_texto_customizado=corpo_texto
                )

        return Response({
            'message': 'Banca concluída com sucesso',
            'banca': BancaFase1Serializer(banca, context={'request': request}).data,
            'etapa_atual': tcc.etapa_atual,
            'etapa_display': tcc.get_etapa_atual_display()
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'], url_path='avaliacoes-fase1', permission_classes=[IsAuthenticated])
    def avaliacoes_fase1(self, request, pk=None):
        """
        GET /api/tccs/{id}/avaliacoes-fase1/
        Lista avaliações da Fase I:
        - Coordenador: vê todas
        - Orientador/Coorientador: vê todas
        - Avaliador: vê apenas a sua própria
        """
        tcc = self.get_object()
        usuario = request.user

        # Filtrar conforme papel
        avaliacoes = AvaliacaoFase1.objects.filter(tcc=tcc)

        if usuario.tipo_usuario == 'COORDENADOR':
            # Coordenador vê todas
            pass
        elif tcc.orientador == usuario or tcc.coorientador == usuario:
            # Orientador/coorientador vê todas
            pass
        else:
            # Avaliador vê apenas a sua
            avaliacoes = avaliacoes.filter(avaliador=usuario)

        serializer = AvaliacaoFase1Serializer(avaliacoes, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='avaliacoes-fase1/enviar', permission_classes=[IsAuthenticated])
    def enviar_avaliacao_fase1(self, request, pk=None):
        """
        POST /api/tccs/{id}/avaliacoes-fase1/enviar/
        Avaliador cria/atualiza sua avaliação:
        - Campos: nota_resumo (0-1.0), nota_introducao (0-2.0), nota_revisao (0-2.0),
                 nota_desenvolvimento (0-3.5), nota_conclusoes (0-1.5), parecer, status (PENDENTE/ENVIADO)
        - Validar que avaliador está na banca
        - Validar que pode editar (não bloqueado)
        """
        from django.db import transaction

        tcc = self.get_object()
        usuario = request.user

        # Verificar se é avaliador deste TCC
        try:
            avaliacao = AvaliacaoFase1.objects.get(tcc=tcc, avaliador=usuario)
        except AvaliacaoFase1.DoesNotExist:
            return Response(
                {'detail': 'Você não é avaliador deste TCC'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Verificar se pode editar
        if avaliacao.status == 'BLOQUEADO':
            return Response(
                {'detail': 'Esta avaliação está bloqueada pelo coordenador'},
                status=status.HTTP_403_FORBIDDEN
            )

        # CASO ESPECIAL: Durante AGUARDANDO_AJUSTES_FINAIS, permitir edição se avaliação está PENDENTE
        # (coordenador solicitou ajustes e desbloqueou esta avaliação específica)
        if tcc.etapa_atual == EtapaTCC.AGUARDANDO_AJUSTES_FINAIS and avaliacao.status == 'PENDENTE':
            # Permitir edição - pular verificações de prazo e flags globais
            pass
        else:
            # Verificações normais de prazo e permissões
            if tcc.avaliacao_fase1_bloqueada:
                return Response(
                    {'detail': 'Avaliações da Fase I foram bloqueadas pelo coordenador'},
                    status=status.HTTP_403_FORBIDDEN
                )

            # Validar prazo e permissões usando calcular_permissoes_tcc
            permissoes = calcular_permissoes_tcc(tcc)
            if not permissoes.get('pode_editar_fase1', False):
                # Definir mensagem apropriada baseada no motivo do bloqueio
                if tcc.avaliacao_fase1_bloqueada:
                    mensagem = 'Avaliações da Fase I foram bloqueadas pelo coordenador'
                elif tcc.liberar_fase1:
                    # Liberação manual estava ativa mas não é suficiente (pode estar bloqueada)
                    mensagem = 'Avaliações da Fase I estão bloqueadas'
                else:
                    mensagem = 'Período de avaliação da Fase I encerrado. Solicite liberação ao coordenador.'

                return Response({'detail': mensagem}, status=status.HTTP_403_FORBIDDEN)

        # Guardar status anterior para criar evento correto
        status_anterior = avaliacao.status

        # Validar dados
        serializer = AvaliacaoFase1EscritaSerializer(avaliacao, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)

        # Usar transação atômica
        with transaction.atomic():
            # Salvar avaliação
            avaliacao_atualizada = serializer.save()

            # Criar evento se foi enviada (PENDENTE → ENVIADO)
            status_novo = request.data.get('status', status_anterior)
            if status_novo == 'ENVIADO' and status_anterior != 'ENVIADO':
                EventoTimeline.objects.create(
                    tcc=tcc,
                    usuario=usuario,
                    tipo_evento=TipoEvento.AVALIACAO_ENVIADA,
                    descricao=f'Avaliação da Fase I enviada por {usuario.nome_completo}',
                    detalhes_json={'avaliacao_id': avaliacao.id},
                    visibilidade=Visibilidade.ORIENTADOR_COORDENADOR
                )

                # Verificar se TODAS as avaliações foram finalizadas (ENVIADO ou BLOQUEADO)
                total_avaliacoes = AvaliacaoFase1.objects.filter(tcc=tcc).count()
                avaliacoes_finalizadas = AvaliacaoFase1.objects.filter(
                    tcc=tcc,
                    status__in=['ENVIADO', 'BLOQUEADO']
                ).count()

                # Se todas foram finalizadas E TCC está em AVALIACAO_FASE_1, mover para VALIDACAO_FASE_1
                if total_avaliacoes > 0 and avaliacoes_finalizadas == total_avaliacoes and tcc.etapa_atual == EtapaTCC.AVALIACAO_FASE_1:
                    tcc.etapa_atual = EtapaTCC.VALIDACAO_FASE_1
                    tcc.save()

                    # Criar evento de transição de etapa
                    EventoTimeline.objects.create(
                        tcc=tcc,
                        usuario=None,  # Sistema
                        tipo_evento=TipoEvento.VALIDACAO_COORDENADOR,
                        descricao=f'Todas as avaliações da Fase I foram enviadas. TCC avançou para etapa: Validação - Fase 1. Aguardando análise do coordenador.',
                        detalhes_json={
                            'etapa_anterior': EtapaTCC.AVALIACAO_FASE_1,
                            'etapa_nova': EtapaTCC.VALIDACAO_FASE_1,
                            'total_avaliacoes': total_avaliacoes
                        },
                        visibilidade=Visibilidade.ORIENTADOR_COORDENADOR
                    )

            elif status_novo == 'PENDENTE' and status_anterior == 'ENVIADO':
                # Cancelamento de envio (reabertura)
                EventoTimeline.objects.create(
                    tcc=tcc,
                    usuario=usuario,
                    tipo_evento=TipoEvento.AVALIACAO_REABERTA,
                    descricao=f'Avaliação da Fase I reaberta por {usuario.nome_completo} para edição',
                    detalhes_json={'avaliacao_id': avaliacao.id},
                    visibilidade=Visibilidade.ORIENTADOR_COORDENADOR
                )

        serializer_response = AvaliacaoFase1Serializer(avaliacao_atualizada, context={'request': request})
        return Response(serializer_response.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='avaliacao-fase1/bloquear', permission_classes=[IsCoordenador])
    def bloquear_avaliacao_fase1(self, request, pk=None):
        """
        POST /api/tccs/{id}/avaliacao-fase1/bloquear/
        Coordenador bloqueia TODAS as avaliações:
        - Seta tcc.avaliacao_fase1_bloqueada = True
        - Muda status de avaliações ENVIADAS para BLOQUEADO
        - Cria evento
        """
        from django.db import transaction

        tcc = self.get_object()

        # Validar que está em AVALIACAO_FASE_1 ou VALIDACAO_FASE_1
        if tcc.etapa_atual not in [EtapaTCC.AVALIACAO_FASE_1, EtapaTCC.VALIDACAO_FASE_1]:
            return Response(
                {'detail': f'TCC deve estar em AVALIACAO_FASE_1 ou VALIDACAO_FASE_1. Etapa atual: {tcc.get_etapa_atual_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar se já está bloqueado
        if tcc.avaliacao_fase1_bloqueada:
            return Response(
                {'detail': 'Avaliações já estão bloqueadas'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Usar transação atômica
        with transaction.atomic():
            # Bloquear TCC
            tcc.avaliacao_fase1_bloqueada = True
            tcc.save()

            # Atualizar avaliações ENVIADAS para BLOQUEADO
            avaliacoes_enviadas = AvaliacaoFase1.objects.filter(tcc=tcc, status='ENVIADO')
            count = avaliacoes_enviadas.update(status='BLOQUEADO')

            # Criar evento
            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=request.user,
                tipo_evento=TipoEvento.BLOQUEIO_AVALIACOES,
                descricao=f'Avaliações da Fase I bloqueadas por {request.user.nome_completo} ({count} avaliações bloqueadas)',
                detalhes_json={'avaliacoes_bloqueadas': count},
                visibilidade=Visibilidade.ORIENTADOR_COORDENADOR
            )

        return Response({
            'message': 'Avaliações bloqueadas com sucesso',
            'avaliacoes_bloqueadas': count
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='avaliacao-fase1/desbloquear', permission_classes=[IsCoordenador])
    def desbloquear_avaliacao_fase1(self, request, pk=None):
        """
        POST /api/tccs/{id}/avaliacao-fase1/desbloquear/
        Coordenador desbloqueia TODAS as avaliações:
        - Seta tcc.avaliacao_fase1_bloqueada = False
        - Muda status de avaliações BLOQUEADO para ENVIADO
        - Cria evento
        """
        from django.db import transaction

        tcc = self.get_object()

        # Verificar se está bloqueado
        if not tcc.avaliacao_fase1_bloqueada:
            return Response(
                {'detail': 'Avaliações não estão bloqueadas'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Usar transação atômica
        with transaction.atomic():
            # Desbloquear TCC
            tcc.avaliacao_fase1_bloqueada = False
            tcc.save()

            # Atualizar avaliações BLOQUEADO para ENVIADO
            avaliacoes_bloqueadas = AvaliacaoFase1.objects.filter(tcc=tcc, status='BLOQUEADO')
            count = avaliacoes_bloqueadas.update(status='ENVIADO')

            # Criar evento
            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=request.user,
                tipo_evento=TipoEvento.DESBLOQUEIO_AVALIACOES,
                descricao=f'Avaliações da Fase I desbloqueadas por {request.user.nome_completo} ({count} avaliações desbloqueadas)',
                detalhes_json={'avaliacoes_desbloqueadas': count},
                visibilidade=Visibilidade.ORIENTADOR_COORDENADOR
            )

        return Response({
            'message': 'Avaliações desbloqueadas com sucesso',
            'avaliacoes_desbloqueadas': count
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='solicitar-ajustes-fase1', permission_classes=[IsCoordenador])
    def solicitar_ajustes_fase1(self, request, pk=None):
        """
        POST /api/tccs/{id}/solicitar-ajustes-fase1/
        Coordenador solicita ajustes a avaliadores específicos:
        - Campos: avaliadores (lista de IDs), mensagem (opcional)
        - Muda status de BLOQUEADO para PENDENTE para os selecionados
        - Cria evento
        """
        from django.db import transaction

        tcc = self.get_object()

        # Validar dados
        avaliadores_ids = request.data.get('avaliadores', [])
        mensagem = request.data.get('mensagem', '')

        if not avaliadores_ids:
            return Response(
                {'detail': 'Campo avaliadores é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not isinstance(avaliadores_ids, list):
            return Response(
                {'detail': 'Campo avaliadores deve ser uma lista de IDs'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar que avaliações existem
        avaliacoes = AvaliacaoFase1.objects.filter(tcc=tcc, avaliador__id__in=avaliadores_ids)
        if avaliacoes.count() != len(avaliadores_ids):
            return Response(
                {'detail': 'Alguns avaliadores não possuem avaliação neste TCC'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Usar transação atômica
        with transaction.atomic():
            # Reabrir avaliações (BLOQUEADO → PENDENTE e limpar enviado_em)
            avaliacoes_bloqueadas = avaliacoes.filter(status='BLOQUEADO')
            count = avaliacoes_bloqueadas.count()

            # Atualizar cada avaliação individualmente para limpar enviado_em
            for avaliacao in avaliacoes_bloqueadas:
                avaliacao.status = 'PENDENTE'
                avaliacao.enviado_em = None
                avaliacao.save()

            # Desbloquear TCC
            tcc.avaliacao_fase1_bloqueada = False

            # Verificar se TODAS as avaliações estão PENDENTE agora (reabriu tudo)
            total_avaliacoes = AvaliacaoFase1.objects.filter(tcc=tcc).count()
            avaliacoes_pendentes = AvaliacaoFase1.objects.filter(tcc=tcc, status='PENDENTE').count()

            # Se TODAS as avaliações são PENDENTE E estava em VALIDACAO_FASE_1, voltar para AVALIACAO_FASE_1
            # Caso contrário, TCC permanece em VALIDACAO_FASE_1 (solicitação parcial)
            etapa_mudou = False
            if avaliacoes_pendentes == total_avaliacoes and tcc.etapa_atual == EtapaTCC.VALIDACAO_FASE_1:
                tcc.etapa_atual = EtapaTCC.AVALIACAO_FASE_1
                etapa_mudou = True

            tcc.save()

            # Criar evento
            nomes_avaliadores = ', '.join([av.avaliador.nome_completo for av in avaliacoes])
            descricao = f'Ajustes solicitados por {request.user.nome_completo} para: {nomes_avaliadores}'
            if mensagem:
                descricao += f'. Mensagem: {mensagem}'

            if etapa_mudou:
                descricao += '. TCC retornou para etapa: Avaliação - Fase 1'
            else:
                descricao += '. TCC permanece em: Validação - Fase 1 (ajustes parciais)'

            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=request.user,
                tipo_evento=TipoEvento.SOLICITACAO_AJUSTES,
                descricao=descricao,
                detalhes_json={
                    'avaliadores': avaliadores_ids,
                    'mensagem': mensagem,
                    'etapa_anterior': EtapaTCC.VALIDACAO_FASE_1,
                    'etapa_nova': EtapaTCC.AVALIACAO_FASE_1
                },
                visibilidade=Visibilidade.ORIENTADOR_COORDENADOR
            )

            # Notificar cada avaliador sobre os ajustes solicitados COM e-mail
            from notificacoes.services import criar_notificacao_com_email
            from notificacoes.constants import TipoNotificacao, PrioridadeNotificacao

            for avaliacao in avaliacoes:
                notif_mensagem = f'O coordenador {request.user.nome_completo} solicitou ajustes na sua avaliação da Fase I do TCC "{tcc.titulo}".'
                if mensagem:
                    notif_mensagem += f' Mensagem: {mensagem}'

                criar_notificacao_com_email(
                    usuario=avaliacao.avaliador,
                    tipo=TipoNotificacao.AJUSTES_SOLICITADOS,
                    titulo='Ajustes Solicitados na Avaliação - Fase I',
                    mensagem=notif_mensagem,
                    campo_preferencia='prof_resultado_fase_1',  # Reutilizar preferência de Fase I
                    action_url=f'/tccs/{tcc.id}',
                    tcc_id=tcc.id,
                    prioridade=PrioridadeNotificacao.ALTA
                )

        return Response({
            'message': 'Ajustes solicitados com sucesso',
            'avaliacoes_reabertas': count,
            'avaliadores': avaliadores_ids,
            'tcc_desbloqueado': not tcc.avaliacao_fase1_bloqueada
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='aprovar-avaliacoes-fase1', permission_classes=[IsCoordenador])
    def aprovar_avaliacoes_fase1(self, request, pk=None):
        """
        POST /api/tccs/{id}/aprovar-avaliacoes-fase1/
        Coordenador aprova avaliações (parcial ou completa):

        APROVAÇÃO PARCIAL (com campo 'avaliadores'):
        - Bloqueia apenas avaliações dos IDs especificados
        - TCC permanece em AVALIACAO_FASE_1
        - Cria evento

        APROVAÇÃO COMPLETA (sem campo 'avaliadores'):
        - Calcula NF1 = média de todas as notas finais
        - Bloqueia todas as avaliações
        - Move TCC para AGENDAMENTO_APRESENTACAO (se NF1 ≥ 6) ou REPROVADO_FASE_1 (se < 6)
        - Cria evento com resultado
        """
        from django.db import transaction
        from decimal import Decimal

        tcc = self.get_object()

        # Validar etapa - aceita tanto AVALIACAO_FASE_1 quanto VALIDACAO_FASE_1
        if tcc.etapa_atual not in [EtapaTCC.AVALIACAO_FASE_1, EtapaTCC.VALIDACAO_FASE_1]:
            return Response(
                {'detail': f'TCC deve estar em AVALIACAO_FASE_1 ou VALIDACAO_FASE_1. Etapa atual: {tcc.get_etapa_atual_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        avaliadores_ids = request.data.get('avaliadores', None)

        # Usar transação atômica
        with transaction.atomic():
            if avaliadores_ids is not None:
                # APROVAÇÃO PARCIAL
                if not isinstance(avaliadores_ids, list) or not avaliadores_ids:
                    return Response(
                        {'detail': 'Campo avaliadores deve ser uma lista não vazia'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Obter avaliações
                avaliacoes = AvaliacaoFase1.objects.filter(tcc=tcc, avaliador__id__in=avaliadores_ids, status='ENVIADO')

                if not avaliacoes.exists():
                    return Response(
                        {'detail': 'Nenhuma avaliação ENVIADA encontrada para os avaliadores especificados'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Bloquear avaliações selecionadas
                count = avaliacoes.update(status='BLOQUEADO')

                # Criar evento
                nomes_avaliadores = ', '.join([av.avaliador.nome_completo for av in avaliacoes])
                EventoTimeline.objects.create(
                    tcc=tcc,
                    usuario=request.user,
                    tipo_evento=TipoEvento.APROVACAO_PARCIAL,
                    descricao=f'Aprovação parcial por {request.user.nome_completo}. Avaliações aprovadas: {nomes_avaliadores}',
                    detalhes_json={'avaliadores': avaliadores_ids},
                    visibilidade=Visibilidade.ORIENTADOR_COORDENADOR
                )

                # Criar notificações para os avaliadores aprovados
                for avaliacao in avaliacoes:
                    criar_notificacao(
                        usuario=avaliacao.avaliador,
                        tipo=TipoNotificacao.AVALIACAO_APROVADA,
                        titulo='Avaliação Fase I Aprovada',
                        mensagem=f'Sua avaliação do TCC "{tcc.titulo}" foi aprovada pelo coordenador.',
                        action_url=f'/tccs/{tcc.id}',
                        tcc_id=tcc.id,
                        prioridade=PrioridadeNotificacao.NORMAL
                    )

                return Response({
                    'message': 'Aprovação parcial concluída',
                    'tipo': 'parcial',
                    'avaliacoes_aprovadas': count
                }, status=status.HTTP_200_OK)

            else:
                # APROVAÇÃO COMPLETA
                # Verificar se todas as avaliações foram enviadas ou bloqueadas
                total_avaliacoes = AvaliacaoFase1.objects.filter(tcc=tcc).count()
                avaliacoes_prontas = AvaliacaoFase1.objects.filter(tcc=tcc, status__in=['ENVIADO', 'BLOQUEADO'])

                if avaliacoes_prontas.count() != total_avaliacoes:
                    pendentes = total_avaliacoes - avaliacoes_prontas.count()
                    return Response(
                        {'detail': f'Existem {pendentes} avaliação(ões) pendente(s). Todas devem estar ENVIADAS ou BLOQUEADAS para aprovação completa'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Calcular NF1 (média das notas finais de cada avaliador)
                notas_finais = []
                for avaliacao in avaliacoes_prontas:
                    # Verificar se todas as notas foram preenchidas
                    if (avaliacao.nota_resumo is None or avaliacao.nota_introducao is None or
                        avaliacao.nota_revisao is None or avaliacao.nota_desenvolvimento is None or
                        avaliacao.nota_conclusoes is None):
                        return Response(
                            {'detail': f'Avaliador {avaliacao.avaliador.nome_completo} não preencheu todas as notas'},
                            status=status.HTTP_400_BAD_REQUEST
                        )

                    # Calcular nota final do avaliador (soma dos 5 critérios)
                    nota_final = avaliacao.calcular_nota_total()
                    notas_finais.append(nota_final)

                # Média das notas finais
                nf1 = sum(notas_finais) / len(notas_finais)
                nf1_arredondado = round(nf1, 2)

                # Bloquear todas as avaliações que ainda estão ENVIADO
                avaliacoes_prontas.filter(status='ENVIADO').update(status='BLOQUEADO')

                # Salvar NF1 no TCC
                tcc.nf1 = Decimal(str(nf1_arredondado))

                # Definir próxima etapa
                if nf1_arredondado >= 6:
                    tcc.etapa_atual = EtapaTCC.AGENDAMENTO_APRESENTACAO
                    resultado = 'APROVADO'
                else:
                    tcc.etapa_atual = EtapaTCC.REPROVADO_FASE_1
                    resultado = 'REPROVADO'

                tcc.save()

                # Criar evento
                EventoTimeline.objects.create(
                    tcc=tcc,
                    usuario=request.user,
                    tipo_evento=TipoEvento.RESULTADO_FASE_1,
                    descricao=f'Resultado da Fase I: {resultado} (NF1 = {nf1_arredondado}). Aprovado por {request.user.nome_completo}. Etapa: {tcc.get_etapa_atual_display()}',
                    detalhes_json={
                        'nf1': str(nf1_arredondado),
                        'resultado': resultado,
                        'notas_finais': [str(n) for n in notas_finais]
                    },
                    visibilidade=Visibilidade.TODOS
                )

                # Notificar orientador (e co-orientador) sobre resultado da Fase I
                from notificacoes.services import criar_notificacao_com_email

                rodape_html = """<p>---<br>
Portal TCC<br>
Esta é uma notificação automática. Para mais informações, acesse o sistema.</p>"""

                rodape_texto = """---
Portal TCC
Esta é uma notificação automática. Para mais informações, acesse o sistema."""

                if resultado == 'APROVADO':
                    def _gerar_email_fase1_aprovado(nome_destinatario):
                        corpo_html = f"""
<p>Olá, {nome_destinatario},</p>

<p>A avaliação da banca da <b>Fase I</b> do TCC foi concluída e confirmada pela coordenação.</p>

<p>O aluno <b>{tcc.aluno.nome_completo}</b> foi <b>aprovado</b> na Fase I.</p>

<p><b>Tema:</b> {tcc.titulo}</p>

<p>O agendamento da defesa já pode ser realizado no sistema.</p>

<p>Acesse o sistema: <a href="http://localhost:5500">http://localhost:5500</a></p>

{rodape_html}
"""
                        corpo_texto = f"""Olá, {nome_destinatario},

A avaliação da banca da Fase I do TCC foi concluída e confirmada pela coordenação.

O aluno {tcc.aluno.nome_completo} foi APROVADO na Fase I.

Tema: {tcc.titulo}

O agendamento da defesa já pode ser realizado no sistema.

Acesse o sistema: http://localhost:5500

{rodape_texto}
"""
                        return corpo_html, corpo_texto

                    html_ori, texto_ori = _gerar_email_fase1_aprovado(tcc.orientador.nome_completo)
                    criar_notificacao_com_email(
                        usuario=tcc.orientador,
                        tipo=TipoNotificacao.RESULTADO_FASE_1,
                        titulo='Resultado Fase I: Aprovado',
                        mensagem=f'A avaliação da Fase I do TCC "{tcc.titulo}" foi aprovada. O agendamento da defesa já pode ser realizado.',
                        campo_preferencia='prof_resultado_fase_1',
                        action_url=f'/tccs/{tcc.id}',
                        tcc_id=tcc.id,
                        prioridade=PrioridadeNotificacao.ALTA,
                        corpo_html_customizado=html_ori,
                        corpo_texto_customizado=texto_ori
                    )

                    if tcc.coorientador:
                        html_coori, texto_coori = _gerar_email_fase1_aprovado(tcc.coorientador.nome_completo)
                        criar_notificacao_com_email(
                            usuario=tcc.coorientador,
                            tipo=TipoNotificacao.RESULTADO_FASE_1,
                            titulo='Resultado Fase I: Aprovado',
                            mensagem=f'A avaliação da Fase I do TCC "{tcc.titulo}" foi aprovada. O agendamento da defesa já pode ser realizado.',
                            campo_preferencia=_pref(tcc.coorientador, 'prof_resultado_fase_1'),
                            action_url=f'/tccs/{tcc.id}',
                            tcc_id=tcc.id,
                            prioridade=PrioridadeNotificacao.ALTA,
                            corpo_html_customizado=html_coori,
                            corpo_texto_customizado=texto_coori
                        )

                else:
                    # REPROVADO na Fase I
                    def _gerar_email_fase1_reprovado(nome_destinatario):
                        corpo_html = f"""
<p>Olá, {nome_destinatario},</p>

<p>A avaliação da banca da <b>Fase I</b> do TCC foi concluída e confirmada pela coordenação.</p>

<p>O aluno <b>{tcc.aluno.nome_completo}</b> foi <b>reprovado</b> na Fase I.</p>

<p><b>Tema:</b> {tcc.titulo}</p>

{rodape_html}
"""
                        corpo_texto = f"""Olá, {nome_destinatario},

A avaliação da banca da Fase I do TCC foi concluída e confirmada pela coordenação.

O aluno {tcc.aluno.nome_completo} foi REPROVADO na Fase I.

Tema: {tcc.titulo}

{rodape_texto}
"""
                        return corpo_html, corpo_texto

                    html_ori, texto_ori = _gerar_email_fase1_reprovado(tcc.orientador.nome_completo)
                    criar_notificacao_com_email(
                        usuario=tcc.orientador,
                        tipo=TipoNotificacao.RESULTADO_FASE_1,
                        titulo='Resultado Fase I: Reprovado',
                        mensagem=f'O aluno {tcc.aluno.nome_completo} foi reprovado na Fase I do TCC "{tcc.titulo}".',
                        campo_preferencia='prof_resultado_fase_1',
                        action_url=f'/tccs/{tcc.id}',
                        tcc_id=tcc.id,
                        prioridade=PrioridadeNotificacao.ALTA,
                        corpo_html_customizado=html_ori,
                        corpo_texto_customizado=texto_ori
                    )

                    if tcc.coorientador:
                        html_coori, texto_coori = _gerar_email_fase1_reprovado(tcc.coorientador.nome_completo)
                        criar_notificacao_com_email(
                            usuario=tcc.coorientador,
                            tipo=TipoNotificacao.RESULTADO_FASE_1,
                            titulo='Resultado Fase I: Reprovado',
                            mensagem=f'O aluno {tcc.aluno.nome_completo} foi reprovado na Fase I do TCC "{tcc.titulo}".',
                            campo_preferencia=_pref(tcc.coorientador, 'prof_resultado_fase_1'),
                            action_url=f'/tccs/{tcc.id}',
                            tcc_id=tcc.id,
                            prioridade=PrioridadeNotificacao.ALTA,
                            corpo_html_customizado=html_coori,
                            corpo_texto_customizado=texto_coori
                        )

                return Response({
                    'message': 'Aprovação completa concluída',
                    'tipo': 'completa',
                    'nf1': nf1_arredondado,
                    'resultado': resultado,
                    'etapa_atual': tcc.etapa_atual,
                    'etapa_display': tcc.get_etapa_atual_display()
                }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'])
    def timeline(self, request, pk=None):
        """
        GET /api/tccs/{id}/timeline/
        Retorna eventos deste TCC (rota aninhada).
        """
        tcc = self.get_object()
        usuario = request.user

        # Filtrar eventos por visibilidade
        from .constants import Visibilidade as Vis

        if usuario.tipo_usuario == 'COORDENADOR':
            eventos = EventoTimeline.objects.filter(tcc=tcc)
        elif usuario.tipo_usuario == 'ALUNO' and tcc.aluno == usuario:
            eventos = EventoTimeline.objects.filter(tcc=tcc, visibilidade=Vis.TODOS)
        elif usuario.tipo_usuario in ['PROFESSOR', 'COORDENADOR']:
            if tcc.orientador == usuario or tcc.coorientador == usuario:
                eventos = EventoTimeline.objects.filter(
                    tcc=tcc
                ).filter(
                    Q(visibilidade=Vis.TODOS) | Q(visibilidade=Vis.ORIENTADOR_COORDENADOR)
                )
            else:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied('Você não tem permissão para acessar este TCC')
        else:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Você não tem permissão para acessar este TCC')

        eventos = eventos.order_by('-timestamp')
        serializer = EventoTimelineSerializer(eventos, many=True, context={'request': request})
        return Response(serializer.data)

    # Endpoints da Fase II

    @action(detail=True, methods=['get', 'post', 'put'], url_path='agendamento-defesa', permission_classes=[IsProfessorOrCoordenador])
    def agendar_defesa(self, request, pk=None):
        """
        GET/POST/PUT /api/tccs/{id}/agendamento-defesa/
        Orientador agenda ou edita a defesa da Fase II.
        - GET: Retorna agendamento existente (se houver)
        - POST/PUT: Cria ou atualiza agendamento (data, hora, local)
        """
        from django.db import transaction
        from .models import AgendamentoDefesa
        from .serializers import AgendamentoDefesaSerializer

        tcc = self.get_object()
        usuario = request.user

        # Validar que usuário é orientador ou coorientador
        if tcc.orientador != usuario and tcc.coorientador != usuario:
            return Response(
                {'detail': 'Apenas orientador ou coorientador pode agendar defesa'},
                status=status.HTTP_403_FORBIDDEN
            )

        # GET: Retornar agendamento existente
        if request.method == 'GET':
            try:
                agendamento = AgendamentoDefesa.objects.get(tcc=tcc)
                serializer = AgendamentoDefesaSerializer(agendamento, context={'request': request})
                return Response(serializer.data)
            except AgendamentoDefesa.DoesNotExist:
                return Response({'detail': 'Agendamento ainda não foi criado'}, status=status.HTTP_404_NOT_FOUND)

        # POST/PUT: Criar ou atualizar agendamento
        # Validar condições para criar/editar:
        # 1. Se agendamento não existe, TCC deve estar em AGENDAMENTO_APRESENTACAO
        # 2. Se agendamento existe (edição), permitir editar enquanto:
        #    - Todas as avaliações Fase II NÃO estão bloqueadas (ainda pode mudar)
        #    - OU data/hora da defesa ainda não passou (prazo não expirou)

        agendamento_existe = AgendamentoDefesa.objects.filter(tcc=tcc).exists()

        if not agendamento_existe:
            # Criação: exige AGENDAMENTO_APRESENTACAO
            if tcc.etapa_atual != EtapaTCC.AGENDAMENTO_APRESENTACAO:
                return Response(
                    {'detail': f'Para criar agendamento, TCC deve estar em AGENDAMENTO_APRESENTACAO. Etapa atual: {tcc.get_etapa_atual_display()}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            # Edição: verificar se pode editar
            from .models import AvaliacaoFase2
            from datetime import datetime

            # Verificar se todas as avaliações Fase II estão bloqueadas
            todas_bloqueadas = AvaliacaoFase2.objects.filter(tcc=tcc, status='BLOQUEADO').count() == 3

            # Verificar se prazo expirou (data/hora da defesa já passou)
            agendamento_atual = AgendamentoDefesa.objects.get(tcc=tcc)
            data_hora_defesa = datetime.combine(agendamento_atual.data, agendamento_atual.hora)
            data_hora_defesa = timezone.make_aware(data_hora_defesa)
            prazo_expirado = data_hora_defesa < timezone.now()

            # Se todas bloqueadas E prazo expirou, não pode editar
            if todas_bloqueadas and prazo_expirado:
                return Response(
                    {'detail': 'Não é possível editar: todas as avaliações estão bloqueadas e o prazo expirou'},
                    status=status.HTTP_403_FORBIDDEN
                )

        # Verificar se já existe agendamento
        try:
            agendamento = AgendamentoDefesa.objects.get(tcc=tcc)
            # Se existe, é uma atualização (PUT)
            serializer = AgendamentoDefesaSerializer(agendamento, data=request.data, partial=True, context={'request': request})
        except AgendamentoDefesa.DoesNotExist:
            # Se não existe, é uma criação (POST)
            serializer = AgendamentoDefesaSerializer(data=request.data, context={'request': request})

        serializer.is_valid(raise_exception=True)

        # Usar transação atômica
        with transaction.atomic():
            # Salvar agendamento
            is_new = False
            try:
                AgendamentoDefesa.objects.get(tcc=tcc)
            except AgendamentoDefesa.DoesNotExist:
                is_new = True

            if 'tcc' not in serializer.validated_data:
                agendamento = serializer.save(tcc=tcc, agendado_por=usuario)
            else:
                agendamento = serializer.save(agendado_por=usuario)

            # Se é novo agendamento, mover TCC para APRESENTACAO_FASE_2 e criar avaliações
            if is_new and request.method == 'POST':
                # Mover TCC para APRESENTACAO_FASE_2
                tcc.etapa_atual = EtapaTCC.APRESENTACAO_FASE_2
                tcc.save()

                # Criar avaliações da Fase II para todos os membros da banca
                try:
                    banca = BancaFase1.objects.get(tcc=tcc)
                    membros = MembroBanca.objects.filter(banca=banca)

                    from .models import AvaliacaoFase2
                    for membro in membros:
                        AvaliacaoFase2.objects.get_or_create(
                            tcc=tcc,
                            avaliador=membro.usuario,
                            defaults={'status': 'PENDENTE'}
                        )
                except BancaFase1.DoesNotExist:
                    pass  # Banca não foi criada ainda

            # Criar evento
            data_formatada = agendamento.data.strftime('%d/%m/%Y')
            hora_formatada = agendamento.hora.strftime('%H:%M')

            if is_new:
                descricao = f'Defesa agendada por {usuario.nome_completo} para {data_formatada} às {hora_formatada}. Local: {agendamento.local}. Etapa: {tcc.get_etapa_atual_display()}'
            else:
                descricao = f'Agendamento editado por {usuario.nome_completo}. Nova data: {data_formatada} às {hora_formatada}. Local: {agendamento.local}'

            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=usuario,
                tipo_evento=TipoEvento.AGENDAMENTO_DEFESA,
                descricao=descricao,
                detalhes_json={
                    'agendamento_id': agendamento.id,
                    'data': agendamento.data.isoformat(),
                    'hora': agendamento.hora.isoformat(),
                    'local': agendamento.local
                },
                visibilidade=Visibilidade.TODOS
            )

            # Criar notificações COM e-mail
            if is_new:  # Apenas quando for novo agendamento (não edição)
                from notificacoes.services import criar_notificacao_com_email, criar_notificacao_em_massa_com_email

                # Notificar aluno
                criar_notificacao_com_email(
                    usuario=tcc.aluno,
                    tipo=TipoNotificacao.DEFESA_AGENDADA,
                    titulo='Defesa Agendada',
                    mensagem=f'Sua defesa foi agendada para {data_formatada} às {hora_formatada}. Local: {agendamento.local}',
                    campo_preferencia='aluno_agendamento_defesa',
                    action_url=f'/tccs/{tcc.id}',
                    tcc_id=tcc.id,
                    prioridade=PrioridadeNotificacao.ALTA
                )

                # Notificar coordenador
                from users.models import Usuario
                coordenadores = Usuario.objects.filter(tipo_usuario='COORDENADOR')
                criar_notificacao_em_massa_com_email(
                    usuarios=list(coordenadores),
                    tipo=TipoNotificacao.DEFESA_AGENDADA,
                    titulo='Defesa Agendada',
                    mensagem=f'Defesa do TCC "{tcc.titulo}" foi agendada para {data_formatada} às {hora_formatada}.',
                    campo_preferencia='coord_defesa_agendada',
                    action_url=f'/tccs/{tcc.id}',
                    tcc_id=tcc.id,
                    prioridade=PrioridadeNotificacao.NORMAL
                )

        return Response(
            AgendamentoDefesaSerializer(agendamento, context={'request': request}).data,
            status=status.HTTP_201_CREATED if request.method == 'POST' else status.HTTP_200_OK
        )


    @action(detail=True, methods=['get'], url_path='avaliacoes-fase2', permission_classes=[IsAuthenticated])
    def avaliacoes_fase2(self, request, pk=None):
        """
        GET /api/tccs/{id}/avaliacoes-fase2/
        Lista avaliações da Fase II (apresentação):
        - Coordenador: vê todas
        - Orientador/Coorientador: vê todas
        - Avaliador: vê apenas a sua própria
        """
        from .models import AvaliacaoFase2
        from .serializers import AvaliacaoFase2Serializer

        tcc = self.get_object()
        usuario = request.user

        # Filtrar conforme papel
        avaliacoes = AvaliacaoFase2.objects.filter(tcc=tcc)

        if usuario.tipo_usuario == 'COORDENADOR':
            # Coordenador vê todas
            pass
        elif tcc.orientador == usuario or tcc.coorientador == usuario:
            # Orientador/coorientador vê todas
            pass
        else:
            # Avaliador vê apenas a sua
            avaliacoes = avaliacoes.filter(avaliador=usuario)

        serializer = AvaliacaoFase2Serializer(avaliacoes, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='avaliacoes-fase2/enviar', permission_classes=[IsAuthenticated])
    def enviar_avaliacao_fase2(self, request, pk=None):
        """
        POST /api/tccs/{id}/avaliacoes-fase2/enviar/
        Avaliador cria/atualiza sua avaliação da apresentação:
        - Campos: 5 notas da Fase II, parecer, status (PENDENTE/ENVIADO)
        - Validar que avaliador está na banca
        - Validar que pode editar (defesa já aconteceu)
        """
        from django.db import transaction
        from .models import AvaliacaoFase2, MembroBanca
        from .serializers import AvaliacaoFase2Serializer, AvaliacaoFase2EscritaSerializer

        tcc = self.get_object()
        usuario = request.user

        # Verificar se é avaliador deste TCC (membro da banca)
        eh_membro_banca = MembroBanca.objects.filter(
            banca__tcc=tcc,
            usuario=usuario
        ).exists()

        if not eh_membro_banca:
            return Response(
                {'detail': 'Você não é membro da banca deste TCC'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Validar que TCC está em APRESENTACAO_FASE_2, ANALISE_FINAL_COORDENADOR ou AGUARDANDO_AJUSTES_FINAIS
        # (permite cancelar mesmo após transição automática, pois signal vai reverter)
        # AGUARDANDO_AJUSTES_FINAIS permite que avaliadores selecionados reenviem
        if tcc.etapa_atual not in [EtapaTCC.APRESENTACAO_FASE_2, EtapaTCC.ANALISE_FINAL_COORDENADOR, EtapaTCC.AGUARDANDO_AJUSTES_FINAIS]:
            return Response(
                {'detail': f'TCC deve estar na etapa APRESENTACAO_FASE_2, ANALISE_FINAL_COORDENADOR ou AGUARDANDO_AJUSTES_FINAIS. Etapa atual: {tcc.get_etapa_atual_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que a defesa já aconteceu (data/hora já passou)
        try:
            from .models import AgendamentoDefesa
            agendamento = AgendamentoDefesa.objects.get(tcc=tcc)
            from datetime import datetime

            data_hora_defesa = datetime.combine(agendamento.data, agendamento.hora)
            data_hora_defesa = timezone.make_aware(data_hora_defesa)

            if data_hora_defesa > timezone.now():
                return Response(
                    {'detail': 'A defesa ainda não aconteceu. Avaliação disponível após a apresentação.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        except AgendamentoDefesa.DoesNotExist:
            return Response(
                {'detail': 'Defesa não foi agendada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obter ou criar avaliação
        avaliacao, created = AvaliacaoFase2.objects.get_or_create(
            tcc=tcc,
            avaliador=usuario,
            defaults={'status': 'PENDENTE'}
        )

        # Verificar se pode editar
        if avaliacao.status == 'BLOQUEADO':
            return Response(
                {'detail': 'Esta avaliação está bloqueada pelo coordenador'},
                status=status.HTTP_403_FORBIDDEN
            )

        if tcc.avaliacao_fase2_bloqueada:
            return Response(
                {'detail': 'Avaliações da Fase II foram bloqueadas pelo coordenador'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Guardar status anterior para criar evento correto
        status_anterior = avaliacao.status

        # Validar dados
        serializer = AvaliacaoFase2EscritaSerializer(avaliacao, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)

        # Usar transação atômica
        with transaction.atomic():
            # Salvar avaliação
            avaliacao_atualizada = serializer.save()

            # Criar evento se foi enviada (PENDENTE → ENVIADO)
            status_novo = request.data.get('status', status_anterior)
            if status_novo == 'ENVIADO' and status_anterior != 'ENVIADO':
                EventoTimeline.objects.create(
                    tcc=tcc,
                    usuario=usuario,
                    tipo_evento=TipoEvento.AVALIACAO_ENVIADA,
                    descricao=f'Avaliação da Fase II (apresentação) enviada por {usuario.nome_completo}',
                    detalhes_json={'avaliacao_fase2_id': avaliacao.id},
                    visibilidade=Visibilidade.ORIENTADOR_COORDENADOR
                )

        serializer_response = AvaliacaoFase2Serializer(avaliacao_atualizada, context={'request': request})
        return Response(serializer_response.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='avaliacao-fase2/bloquear', permission_classes=[IsCoordenador])
    def bloquear_avaliacao_fase2(self, request, pk=None):
        """
        POST /api/tccs/{id}/avaliacao-fase2/bloquear/
        Coordenador bloqueia TODAS as avaliações da Fase II:
        - Seta tcc.avaliacao_fase2_bloqueada = True
        - Muda status de avaliações ENVIADAS para BLOQUEADO
        - Cria evento
        """
        from django.db import transaction
        from .models import AvaliacaoFase2

        tcc = self.get_object()

        # Validar que está em ANALISE_FINAL_COORDENADOR
        if tcc.etapa_atual != EtapaTCC.ANALISE_FINAL_COORDENADOR:
            return Response(
                {'detail': f'TCC deve estar em ANALISE_FINAL_COORDENADOR. Etapa atual: {tcc.get_etapa_atual_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar se já está bloqueado
        if tcc.avaliacao_fase2_bloqueada:
            return Response(
                {'detail': 'Avaliações da Fase II já estão bloqueadas'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Usar transação atômica
        with transaction.atomic():
            # Bloquear TCC
            tcc.avaliacao_fase2_bloqueada = True
            tcc.save()

            # Atualizar avaliações ENVIADAS para BLOQUEADO
            avaliacoes_enviadas = AvaliacaoFase2.objects.filter(tcc=tcc, status='ENVIADO')
            count = avaliacoes_enviadas.update(status='BLOQUEADO')

            # Criar evento
            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=request.user,
                tipo_evento=TipoEvento.BLOQUEIO_AVALIACOES,
                descricao=f'Avaliações da Fase II bloqueadas por {request.user.nome_completo} ({count} avaliações bloqueadas)',
                detalhes_json={'avaliacoes_bloqueadas': count, 'fase': 2},
                visibilidade=Visibilidade.ORIENTADOR_COORDENADOR
            )

        return Response({
            'message': 'Avaliações da Fase II bloqueadas com sucesso',
            'avaliacoes_bloqueadas': count
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='avaliacao-fase2/desbloquear', permission_classes=[IsCoordenador])
    def desbloquear_avaliacao_fase2(self, request, pk=None):
        """
        POST /api/tccs/{id}/avaliacao-fase2/desbloquear/
        Coordenador desbloqueia TODAS as avaliações da Fase II:
        - Seta tcc.avaliacao_fase2_bloqueada = False
        - Muda status de avaliações BLOQUEADO para ENVIADO
        - Cria evento
        """
        from django.db import transaction
        from .models import AvaliacaoFase2

        tcc = self.get_object()

        # Verificar se está bloqueado
        if not tcc.avaliacao_fase2_bloqueada:
            return Response(
                {'detail': 'Avaliações da Fase II não estão bloqueadas'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Usar transação atômica
        with transaction.atomic():
            # Desbloquear TCC
            tcc.avaliacao_fase2_bloqueada = False
            tcc.save()

            # Atualizar avaliações BLOQUEADO para ENVIADO
            avaliacoes_bloqueadas = AvaliacaoFase2.objects.filter(tcc=tcc, status='BLOQUEADO')
            count = avaliacoes_bloqueadas.update(status='ENVIADO')

            # Criar evento
            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=request.user,
                tipo_evento=TipoEvento.DESBLOQUEIO_AVALIACOES,
                descricao=f'Avaliações da Fase II desbloqueadas por {request.user.nome_completo} ({count} avaliações desbloqueadas)',
                detalhes_json={'avaliacoes_desbloqueadas': count, 'fase': 2},
                visibilidade=Visibilidade.ORIENTADOR_COORDENADOR
            )

        return Response({
            'message': 'Avaliações da Fase II desbloqueadas com sucesso',
            'avaliacoes_desbloqueadas': count
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='avaliacao-fase2/solicitar-ajustes', permission_classes=[IsCoordenador])
    def solicitar_ajustes_fase2(self, request, pk=None):
        """
        POST /api/tccs/{id}/avaliacao-fase2/solicitar-ajustes/
        Coordenador solicita ajustes em avaliações específicas da Fase II:
        - Recebe lista de IDs de avaliadores + mensagem opcional
        - Desbloqueia apenas as avaliações selecionadas (BLOQUEADO → PENDENTE)
        - TCC continua em ANALISE_FINAL_COORDENADOR
        - Cria evento na timeline
        """
        from django.db import transaction
        from .models import AvaliacaoFase2

        tcc = self.get_object()

        # Validar que TCC está em ANALISE_FINAL_COORDENADOR
        if tcc.etapa_atual != EtapaTCC.ANALISE_FINAL_COORDENADOR:
            return Response(
                {'detail': f'TCC deve estar em ANALISE_FINAL_COORDENADOR. Etapa atual: {tcc.get_etapa_atual_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obter lista de IDs de avaliadores e mensagem
        avaliadores_ids = request.data.get('avaliadores', [])
        mensagem = request.data.get('mensagem', '')

        if not avaliadores_ids:
            return Response(
                {'detail': 'Nenhum avaliador selecionado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que os avaliadores têm avaliações bloqueadas
        avaliacoes = AvaliacaoFase2.objects.filter(
            tcc=tcc,
            avaliador_id__in=avaliadores_ids,
            status='BLOQUEADO'
        )

        if not avaliacoes.exists():
            return Response(
                {'detail': 'Nenhuma avaliação bloqueada encontrada para os avaliadores selecionados'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Usar transação atômica
        with transaction.atomic():
            # Desbloquear apenas as avaliações selecionadas
            count = avaliacoes.update(status='PENDENTE')

            # Criar evento na timeline
            avaliadores_nomes = ', '.join([av.avaliador.nome_completo for av in avaliacoes])
            descricao = f'Ajustes solicitados por {request.user.nome_completo} para {count} avaliador(es): {avaliadores_nomes}'
            if mensagem:
                descricao += f'. Mensagem: "{mensagem}"'

            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=request.user,
                tipo_evento=TipoEvento.SOLICITACAO_AJUSTES,
                descricao=descricao,
                detalhes_json={
                    'avaliadores_ids': avaliadores_ids,
                    'avaliacoes_reabertas': count,
                    'mensagem': mensagem,
                    'fase': 2
                },
                visibilidade=Visibilidade.ORIENTADOR_COORDENADOR
            )

            # Notificar cada avaliador sobre os ajustes solicitados COM e-mail
            from notificacoes.services import criar_notificacao_com_email
            from notificacoes.constants import TipoNotificacao, PrioridadeNotificacao

            for avaliacao in avaliacoes:
                notif_mensagem = f'O coordenador {request.user.nome_completo} solicitou ajustes na sua avaliação da Fase II do TCC "{tcc.titulo}".'
                if mensagem:
                    notif_mensagem += f' Mensagem: {mensagem}'

                criar_notificacao_com_email(
                    usuario=avaliacao.avaliador,
                    tipo=TipoNotificacao.AJUSTES_SOLICITADOS,
                    titulo='Ajustes Solicitados na Avaliação - Fase II',
                    mensagem=notif_mensagem,
                    campo_preferencia='prof_resultado_fase_1',  # Reutilizar preferência (não há específica para Fase II)
                    action_url=f'/tccs/{tcc.id}',
                    tcc_id=tcc.id,
                    prioridade=PrioridadeNotificacao.ALTA
                )

        return Response({
            'message': f'Ajustes solicitados para {count} avaliador(es) com sucesso',
            'avaliacoes_reabertas': count,
            'avaliadores': avaliadores_ids
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='analise-final/solicitar-ajustes-finais', permission_classes=[IsCoordenador])
    def solicitar_ajustes_finais(self, request, pk=None):
        """
        POST /api/tccs/{id}/analise-final/solicitar-ajustes-finais/
        Coordenador solicita ajustes finais em avaliações específicas:
        - Recebe lista de IDs de avaliadores + mensagem opcional
        - Desbloqueia avaliações da Fase II (todos os avaliadores)
        - Desbloqueia avaliações da Fase I (apenas avaliadores externos da banca)
        - Orientador/coorientador: apenas Fase II
        - Muda TCC para AGUARDANDO_AJUSTES_FINAIS
        - Cria evento na timeline
        """
        from django.db import transaction
        from .models import AvaliacaoFase2, AvaliacaoFase1, MembroBanca, BancaFase1

        tcc = self.get_object()

        # Validar que TCC está em ANALISE_FINAL_COORDENADOR
        if tcc.etapa_atual != EtapaTCC.ANALISE_FINAL_COORDENADOR:
            return Response(
                {'detail': f'TCC deve estar em ANALISE_FINAL_COORDENADOR. Etapa atual: {tcc.get_etapa_atual_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obter lista de IDs de avaliadores e mensagem
        avaliadores_ids = request.data.get('avaliadores', [])
        mensagem = request.data.get('mensagem', '')

        if not avaliadores_ids:
            return Response(
                {'detail': 'Nenhum avaliador selecionado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que os avaliadores têm avaliações bloqueadas na Fase II
        avaliacoes_fase2 = AvaliacaoFase2.objects.filter(
            tcc=tcc,
            avaliador_id__in=avaliadores_ids,
            status='BLOQUEADO'
        )

        if not avaliacoes_fase2.exists():
            return Response(
                {'detail': 'Nenhuma avaliação bloqueada encontrada para os avaliadores selecionados'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Usar transação atômica
        with transaction.atomic():
            # Desbloquear avaliações Fase II selecionadas
            count_fase2 = avaliacoes_fase2.update(status='PENDENTE')

            # Para cada avaliador, verificar se é membro da banca (avaliador externo)
            # Se sim, desbloquear também Fase I
            count_fase1 = 0
            avaliadores_com_fase1 = []

            try:
                banca = BancaFase1.objects.get(tcc=tcc)

                for avaliador_id in avaliadores_ids:
                    # Verificar se é orientador ou coorientador
                    eh_orientador = (tcc.orientador_id == avaliador_id)
                    eh_coorientador = (tcc.coorientador_id == avaliador_id) if tcc.coorientador_id else False

                    # Se não é orientador nem coorientador, é avaliador externo da banca
                    if not eh_orientador and not eh_coorientador:
                        # Verificar se tem avaliação Fase I bloqueada
                        avaliacao_fase1 = AvaliacaoFase1.objects.filter(
                            tcc=tcc,
                            avaliador_id=avaliador_id,
                            status='BLOQUEADO'
                        ).first()

                        if avaliacao_fase1:
                            avaliacao_fase1.status = 'PENDENTE'
                            avaliacao_fase1.save()
                            count_fase1 += 1
                            avaliadores_com_fase1.append(avaliador_id)

            except BancaFase1.DoesNotExist:
                pass  # Não há banca criada

            # Mudar TCC para AGUARDANDO_AJUSTES_FINAIS e desbloquear avaliações Fase 2
            # (permite que avaliadores selecionados possam reenviar)
            tcc.etapa_atual = EtapaTCC.AGUARDANDO_AJUSTES_FINAIS
            tcc.avaliacao_fase2_bloqueada = False
            tcc.save()

            # Criar evento na timeline
            avaliadores_nomes = ', '.join([av.avaliador.nome_completo for av in avaliacoes_fase2])
            descricao = f'Ajustes finais solicitados por {request.user.nome_completo} para {len(avaliadores_ids)} avaliador(es): {avaliadores_nomes}'

            # Detalhar quais fases foram desbloqueadas
            if count_fase1 > 0:
                descricao += f'. Fase I desbloqueada para {count_fase1} avaliador(es) externo(s)'

            if mensagem:
                descricao += f'. Mensagem: "{mensagem}"'

            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=request.user,
                tipo_evento=TipoEvento.AJUSTES_SOLICITADOS,
                descricao=descricao,
                detalhes_json={
                    'avaliadores_ids': avaliadores_ids,
                    'avaliacoes_reabertas_fase2': count_fase2,
                    'avaliacoes_reabertas_fase1': count_fase1,
                    'avaliadores_com_fase1': avaliadores_com_fase1,
                    'mensagem': mensagem,
                    'tipo': 'ajustes_finais'
                },
                visibilidade=Visibilidade.ORIENTADOR_COORDENADOR
            )

            # Notificar cada avaliador sobre os ajustes finais COM e-mail
            from notificacoes.services import criar_notificacao_com_email
            from notificacoes.constants import TipoNotificacao, PrioridadeNotificacao

            for avaliacao_fase2 in avaliacoes_fase2:
                avaliador_id = avaliacao_fase2.avaliador_id

                # Construir mensagem específica para cada avaliador
                if avaliador_id in avaliadores_com_fase1:
                    # Avaliador externo - precisa ajustar ambas as fases
                    notif_mensagem = f'O coordenador {request.user.nome_completo} solicitou ajustes finais nas suas avaliações da Fase I e Fase II do TCC "{tcc.titulo}".'
                else:
                    # Orientador/Coorientador - apenas Fase II
                    notif_mensagem = f'O coordenador {request.user.nome_completo} solicitou ajustes finais na sua avaliação da Fase II do TCC "{tcc.titulo}".'

                if mensagem:
                    notif_mensagem += f' Mensagem: {mensagem}'

                criar_notificacao_com_email(
                    usuario=avaliacao_fase2.avaliador,
                    tipo=TipoNotificacao.AJUSTES_SOLICITADOS,
                    titulo='Ajustes Finais Solicitados',
                    mensagem=notif_mensagem,
                    campo_preferencia='prof_resultado_fase_1',  # Reutilizar preferência
                    action_url=f'/tccs/{tcc.id}',
                    tcc_id=tcc.id,
                    prioridade=PrioridadeNotificacao.URGENTE
                )

        return Response({
            'message': f'Ajustes finais solicitados para {len(avaliadores_ids)} avaliador(es). TCC movido para AGUARDANDO_AJUSTES_FINAIS',
            'avaliacoes_reabertas_fase2': count_fase2,
            'avaliacoes_reabertas_fase1': count_fase1,
            'avaliadores_com_fase1': avaliadores_com_fase1,
            'avaliadores': avaliadores_ids,
            'nova_etapa': tcc.etapa_atual
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='analise-final/aprovar-concluir', permission_classes=[IsCoordenador])
    def aprovar_concluir_tcc(self, request, pk=None):
        """
        POST /api/tccs/{id}/analise-final/aprovar-concluir/
        Coordenador aprova e conclui o TCC após análise final:
        - Valida que TCC está em ANALISE_FINAL_COORDENADOR
        - Valida que todas as 3 avaliações da Fase II estão BLOQUEADAS
        - Muda TCC para CONCLUIDO
        - Mantém NF2, MF, resultado já calculados
        - Cria evento na timeline
        """
        from django.db import transaction
        from .models import AvaliacaoFase2

        tcc = self.get_object()

        # Validar que TCC está em ANALISE_FINAL_COORDENADOR
        if tcc.etapa_atual != EtapaTCC.ANALISE_FINAL_COORDENADOR:
            return Response(
                {'detail': f'TCC deve estar em ANALISE_FINAL_COORDENADOR. Etapa atual: {tcc.get_etapa_atual_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que todas as 3 avaliações da Fase II estão BLOQUEADAS
        total_avaliacoes = AvaliacaoFase2.objects.filter(tcc=tcc).count()
        avaliacoes_bloqueadas = AvaliacaoFase2.objects.filter(tcc=tcc, status='BLOQUEADO').count()

        if total_avaliacoes != 3:
            return Response(
                {'detail': f'Devem existir exatamente 3 avaliações da Fase II. Encontradas: {total_avaliacoes}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if avaliacoes_bloqueadas != 3:
            return Response(
                {'detail': f'Todas as 3 avaliações devem estar bloqueadas. Bloqueadas: {avaliacoes_bloqueadas}/3'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Usar transação atômica
        with transaction.atomic():
            # Mudar TCC para CONCLUIDO
            tcc.etapa_atual = EtapaTCC.CONCLUIDO
            tcc.save()

            # Criar evento na timeline
            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=request.user,
                tipo_evento=TipoEvento.CONCLUSAO,
                descricao=f'TCC aprovado e concluído por {request.user.nome_completo}. NF1 = {tcc.nf1}, NF2 = {tcc.nf2}, MF = {tcc.media_final}, Resultado: {tcc.resultado_final}',
                detalhes_json={
                    'nf1': str(tcc.nf1) if tcc.nf1 else None,
                    'nf2': str(tcc.nf2) if tcc.nf2 else None,
                    'media_final': str(tcc.media_final) if tcc.media_final else None,
                    'resultado_final': tcc.resultado_final
                },
                visibilidade=Visibilidade.TODOS
            )

            # Criar notificações
            # Notificar aluno COM e-mail
            from notificacoes.services import criar_notificacao_com_email, criar_notificacao_em_massa_com_email
            criar_notificacao_com_email(
                usuario=tcc.aluno,
                tipo=TipoNotificacao.TCC_CONCLUIDO,
                titulo='TCC Concluído!',
                mensagem=f'Parabéns! Seu TCC "{tcc.titulo}" foi concluído com sucesso!',
                campo_preferencia='aluno_finalizacao_tcc',
                action_url=f'/tccs/{tcc.id}',
                tcc_id=tcc.id,
                prioridade=PrioridadeNotificacao.URGENTE
            )

            # Notificar orientador e coorientador COM e-mail
            for dest in [tcc.orientador, tcc.coorientador]:
                if dest:
                    criar_notificacao_com_email(
                        usuario=dest,
                        tipo=TipoNotificacao.TCC_CONCLUIDO,
                        titulo='TCC Concluído',
                        mensagem=f'O TCC "{tcc.titulo}" foi aprovado e concluído.',
                        campo_preferencia=_pref(dest, 'prof_finalizacao_tcc'),
                        action_url=f'/tccs/{tcc.id}',
                        tcc_id=tcc.id,
                        prioridade=PrioridadeNotificacao.ALTA
                    )

        # Gerar Relatório de Avaliação automaticamente
        try:
            from .relatorio_docx import gerar_relatorio_avaliacao_docx
            from django.core.files.base import ContentFile

            docx_bytes = gerar_relatorio_avaliacao_docx(tcc)
            nome_arquivo = f'Relatorio_Avaliacao_{tcc.aluno.nome_completo.replace(" ", "_")}.docx'

            # Remover relatório anterior se existir
            DocumentoTCC.objects.filter(tcc=tcc, tipo_documento=TipoDocumento.RELATORIO_AVALIACAO).delete()

            doc_relatorio = DocumentoTCC(
                tcc=tcc,
                tipo_documento=TipoDocumento.RELATORIO_AVALIACAO,
                nome_original=nome_arquivo,
                tamanho=len(docx_bytes),
                versao=1,
                enviado_por=request.user,
                status=StatusDocumento.APROVADO,
            )
            doc_relatorio.arquivo.save(nome_arquivo, ContentFile(docx_bytes), save=True)
        except Exception:
            pass  # Não bloquear a conclusão se o relatório falhar

        return Response({
            'message': 'TCC aprovado e concluído com sucesso',
            'etapa_atual': tcc.etapa_atual,
            'etapa_display': tcc.get_etapa_atual_display(),
            'nf1': float(tcc.nf1) if tcc.nf1 else None,
            'nf2': float(tcc.nf2) if tcc.nf2 else None,
            'media_final': float(tcc.media_final) if tcc.media_final else None,
            'resultado_final': tcc.resultado_final
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'], url_path='relatorio-avaliacao', permission_classes=[IsCoordenador])
    def relatorio_avaliacao(self, request, pk=None):
        """
        GET /api/tccs/{id}/relatorio-avaliacao/
        Gera (ou regenera) o Relatório de Avaliação do TCC em DOCX e retorna para download.
        Também salva como DocumentoTCC para ficar disponível em Documentos do TCC.
        """
        from .relatorio_docx import gerar_relatorio_avaliacao_docx
        from django.core.files.base import ContentFile
        from django.http import HttpResponse

        tcc = self.get_object()

        # Validar que o TCC tem avaliações suficientes
        avaliacoes_f1 = AvaliacaoFase1.objects.filter(tcc=tcc, status='BLOQUEADO').count()
        if avaliacoes_f1 == 0:
            return Response(
                {'detail': 'O TCC não possui avaliações da Fase I bloqueadas para gerar o relatório.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            docx_bytes = gerar_relatorio_avaliacao_docx(tcc)
        except Exception as e:
            return Response(
                {'detail': f'Erro ao gerar relatório: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        nome_arquivo = f'Relatorio_Avaliacao_{tcc.aluno.nome_completo.replace(" ", "_")}.docx'

        # Salvar/atualizar como DocumentoTCC
        DocumentoTCC.objects.filter(tcc=tcc, tipo_documento=TipoDocumento.RELATORIO_AVALIACAO).delete()
        doc_relatorio = DocumentoTCC(
            tcc=tcc,
            tipo_documento=TipoDocumento.RELATORIO_AVALIACAO,
            nome_original=nome_arquivo,
            tamanho=len(docx_bytes),
            versao=1,
            enviado_por=request.user,
            status=StatusDocumento.APROVADO,
        )
        doc_relatorio.arquivo.save(nome_arquivo, ContentFile(docx_bytes), save=True)

        # Retornar DOCX para download direto
        response = HttpResponse(
            docx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        return response

    @action(detail=True, methods=['post'], url_path='aprovar-avaliacoes-fase2', permission_classes=[IsCoordenador])
    def aprovar_avaliacoes_fase2(self, request, pk=None):
        """
        POST /api/tccs/{id}/aprovar-avaliacoes-fase2/
        Coordenador aprova avaliações da Fase II:
        - Verifica que todas as avaliações foram enviadas
        - Calcula NF2 = média ponderada de NF1 (40%) + média das notas da apresentação (60%)
        - Bloqueia todas as avaliações
        - Move TCC para APROVADO (se NF2 ≥ 6) ou REPROVADO_FASE_2 (se < 6)
        - Cria evento com resultado
        """
        from django.db import transaction
        from decimal import Decimal
        from .models import AvaliacaoFase2, AvaliacaoFase1

        tcc = self.get_object()

        # Validar etapa
        if tcc.etapa_atual != EtapaTCC.APRESENTACAO_FASE_2:
            return Response(
                {'detail': f'TCC deve estar em APRESENTACAO_FASE_2. Etapa atual: {tcc.get_etapa_atual_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar se todas as avaliações da Fase II foram enviadas
        total_avaliacoes = AvaliacaoFase2.objects.filter(tcc=tcc).count()
        avaliacoes_enviadas = AvaliacaoFase2.objects.filter(tcc=tcc, status='ENVIADO')

        if avaliacoes_enviadas.count() != total_avaliacoes:
            pendentes = total_avaliacoes - avaliacoes_enviadas.count()
            return Response(
                {'detail': f'Existem {pendentes} avaliação(ões) da Fase II pendente(s). Todas devem estar ENVIADAS para aprovação'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar que há pelo menos uma avaliação
        if total_avaliacoes == 0:
            return Response(
                {'detail': 'Não há avaliações da Fase II para aprovar'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Usar transação atômica
        with transaction.atomic():
            # Calcular média das notas da Fase II
            notas_fase2 = []
            for avaliacao in avaliacoes_enviadas:
                # Verificar se todas as notas foram preenchidas
                if (avaliacao.nota_coerencia_conteudo is None or
                    avaliacao.nota_qualidade_apresentacao is None or
                    avaliacao.nota_dominio_tema is None or
                    avaliacao.nota_clareza_fluencia is None or
                    avaliacao.nota_observancia_tempo is None):
                    return Response(
                        {'detail': f'Avaliador {avaliacao.avaliador.nome_completo} não preencheu todas as notas da Fase II'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Calcular nota total do avaliador na Fase II
                nota_total_fase2 = avaliacao.calcular_nota_total()
                notas_fase2.append(nota_total_fase2)

            # Média das notas da Fase II
            media_fase2 = sum(notas_fase2) / len(notas_fase2)

            # Calcular NF1 (média das avaliações da Fase I bloqueadas)
            avaliacoes_fase1 = AvaliacaoFase1.objects.filter(tcc=tcc, status='BLOQUEADO')
            if avaliacoes_fase1.count() == 0:
                return Response(
                    {'detail': 'Não há avaliações da Fase I aprovadas. A aprovação da Fase I deve ser feita primeiro.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            notas_fase1 = []
            for avaliacao in avaliacoes_fase1:
                nota_total_fase1 = avaliacao.calcular_nota_total()
                notas_fase1.append(nota_total_fase1)

            nf1 = sum(notas_fase1) / len(notas_fase1)

            # Calcular NF2: 40% de NF1 + 60% da média da Fase II
            nf2 = (Decimal('0.4') * Decimal(str(nf1))) + (Decimal('0.6') * Decimal(str(media_fase2)))
            nf2_arredondado = round(nf2, 2)

            # Bloquear todas as avaliações da Fase II
            avaliacoes_enviadas.update(status='BLOQUEADO')

            # Definir próxima etapa
            if nf2_arredondado >= 6:
                tcc.etapa_atual = EtapaTCC.APROVADO
                resultado = 'APROVADO'
            else:
                tcc.etapa_atual = EtapaTCC.REPROVADO_FASE_2
                resultado = 'REPROVADO'

            tcc.save()

            # Criar evento
            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=request.user,
                tipo_evento=TipoEvento.RESULTADO_FINAL,
                descricao=f'Resultado Final: {resultado} (NF2 = {nf2_arredondado}). NF1 = {round(nf1, 2)} (40%), Média Apresentação = {round(media_fase2, 2)} (60%). Aprovado por {request.user.nome_completo}. Etapa: {tcc.get_etapa_atual_display()}',
                detalhes_json={
                    'nf2': str(nf2_arredondado),
                    'nf1': str(round(nf1, 2)),
                    'media_fase2': str(round(media_fase2, 2)),
                    'resultado': resultado,
                    'notas_fase2': [str(n) for n in notas_fase2]
                },
                visibilidade=Visibilidade.TODOS
            )

            # Criar notificações
            # Notificar aluno
            criar_notificacao(
                usuario=tcc.aluno,
                tipo=TipoNotificacao.RESULTADO_FINAL,
                titulo=f'Resultado Final: {resultado}',
                mensagem=f'Seu TCC foi {resultado.lower()} com nota final {nf2_arredondado}.',
                action_url=f'/tccs/{tcc.id}',
                tcc_id=tcc.id,
                prioridade=PrioridadeNotificacao.URGENTE
            )

            # Notificar orientador e coorientador com e-mail HTML
            from notificacoes.services import criar_notificacao_com_email

            rodape_html = """<p>---<br>
Portal TCC<br>
Esta é uma notificação automática. Para mais informações, acesse o sistema.</p>"""

            rodape_texto = """---
Portal TCC
Esta é uma notificação automática. Para mais informações, acesse o sistema."""

            if resultado == 'APROVADO':
                def _gerar_email_fase2_aprovado(nome_destinatario):
                    corpo_html = f"""
<p>Olá, {nome_destinatario},</p>

<p>A avaliação da <b>defesa</b> do TCC foi concluída e confirmada pela coordenação.</p>

<p>O aluno <b>{tcc.aluno.nome_completo}</b> foi <b>aprovado</b> na defesa.</p>

<p><b>Tema:</b> {tcc.titulo}</p>

{rodape_html}
"""
                    corpo_texto = f"""Olá, {nome_destinatario},

A avaliação da defesa do TCC foi concluída e confirmada pela coordenação.

O aluno {tcc.aluno.nome_completo} foi APROVADO na defesa.

Tema: {tcc.titulo}

{rodape_texto}
"""
                    return corpo_html, corpo_texto

                html_ori, texto_ori = _gerar_email_fase2_aprovado(tcc.orientador.nome_completo)
                criar_notificacao_com_email(
                    usuario=tcc.orientador,
                    tipo=TipoNotificacao.RESULTADO_FINAL,
                    titulo='Resultado da Defesa: Aprovado',
                    mensagem=f'A avaliação da defesa do TCC "{tcc.titulo}" foi aprovada.',
                    campo_preferencia='prof_resultado_fase_1',
                    action_url=f'/tccs/{tcc.id}',
                    tcc_id=tcc.id,
                    prioridade=PrioridadeNotificacao.ALTA,
                    corpo_html_customizado=html_ori,
                    corpo_texto_customizado=texto_ori
                )

                if tcc.coorientador:
                    html_coori, texto_coori = _gerar_email_fase2_aprovado(tcc.coorientador.nome_completo)
                    criar_notificacao_com_email(
                        usuario=tcc.coorientador,
                        tipo=TipoNotificacao.RESULTADO_FINAL,
                        titulo='Resultado da Defesa: Aprovado',
                        mensagem=f'A avaliação da defesa do TCC "{tcc.titulo}" foi aprovada.',
                        campo_preferencia=_pref(tcc.coorientador, 'prof_resultado_fase_1'),
                        action_url=f'/tccs/{tcc.id}',
                        tcc_id=tcc.id,
                        prioridade=PrioridadeNotificacao.ALTA,
                        corpo_html_customizado=html_coori,
                        corpo_texto_customizado=texto_coori
                    )

            else:
                # REPROVADO na defesa
                def _gerar_email_fase2_reprovado(nome_destinatario):
                    corpo_html = f"""
<p>Olá, {nome_destinatario},</p>

<p>A avaliação da <b>defesa</b> do TCC foi concluída e confirmada pela coordenação.</p>

<p>O aluno <b>{tcc.aluno.nome_completo}</b> foi <b>reprovado</b> na defesa.</p>

<p><b>Tema:</b> {tcc.titulo}</p>

{rodape_html}
"""
                    corpo_texto = f"""Olá, {nome_destinatario},

A avaliação da defesa do TCC foi concluída e confirmada pela coordenação.

O aluno {tcc.aluno.nome_completo} foi REPROVADO na defesa.

Tema: {tcc.titulo}

{rodape_texto}
"""
                    return corpo_html, corpo_texto

                html_ori, texto_ori = _gerar_email_fase2_reprovado(tcc.orientador.nome_completo)
                criar_notificacao_com_email(
                    usuario=tcc.orientador,
                    tipo=TipoNotificacao.RESULTADO_FINAL,
                    titulo='Resultado da Defesa: Reprovado',
                    mensagem=f'O aluno {tcc.aluno.nome_completo} foi reprovado na defesa do TCC "{tcc.titulo}".',
                    campo_preferencia='prof_resultado_fase_1',
                    action_url=f'/tccs/{tcc.id}',
                    tcc_id=tcc.id,
                    prioridade=PrioridadeNotificacao.ALTA,
                    corpo_html_customizado=html_ori,
                    corpo_texto_customizado=texto_ori
                )

                if tcc.coorientador:
                    html_coori, texto_coori = _gerar_email_fase2_reprovado(tcc.coorientador.nome_completo)
                    criar_notificacao_com_email(
                        usuario=tcc.coorientador,
                        tipo=TipoNotificacao.RESULTADO_FINAL,
                        titulo='Resultado da Defesa: Reprovado',
                        mensagem=f'O aluno {tcc.aluno.nome_completo} foi reprovado na defesa do TCC "{tcc.titulo}".',
                        campo_preferencia=_pref(tcc.coorientador, 'prof_resultado_fase_1'),
                        action_url=f'/tccs/{tcc.id}',
                        tcc_id=tcc.id,
                        prioridade=PrioridadeNotificacao.ALTA,
                        corpo_html_customizado=html_coori,
                        corpo_texto_customizado=texto_coori
                    )

            # Notificar membros da banca
            membros_banca = [av.avaliador for av in avaliacoes_enviadas]
            criar_notificacao_em_massa(
                usuarios=membros_banca,
                tipo=TipoNotificacao.RESULTADO_FINAL,
                titulo=f'Resultado Final: {resultado}',
                mensagem=f'TCC "{tcc.titulo}" foi {resultado.lower()} com nota final {nf2_arredondado}.',
                action_url=f'/tccs/{tcc.id}',
                tcc_id=tcc.id,
                prioridade=PrioridadeNotificacao.NORMAL
            )

        return Response({
            'message': 'Aprovação da Fase II concluída',
            'nf1': round(nf1, 2),
            'media_fase2': round(media_fase2, 2),
            'nf2': nf2_arredondado,
            'resultado': resultado,
            'etapa_atual': tcc.etapa_atual,
            'etapa_display': tcc.get_etapa_atual_display()
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='avaliacao-fase1/editar-coordenador', permission_classes=[IsCoordenador])
    def editar_avaliacao_fase1_coordenador(self, request, pk=None):
        """
        POST /api/tccs/{id}/avaliacao-fase1/editar-coordenador/
        Coordenador edita diretamente a avaliação de um avaliador na Fase I.
        Body: { avaliador_id, nota_resumo, nota_introducao, nota_revisao, nota_desenvolvimento, nota_conclusoes, parecer, status }
        """
        from .serializers import AvaliacaoFase1Serializer, AvaliacaoFase1EscritaSerializer

        tcc = self.get_object()
        avaliador_id = request.data.get('avaliador_id')

        if not avaliador_id:
            return Response({'detail': 'avaliador_id é obrigatório'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            avaliacao = AvaliacaoFase1.objects.get(tcc=tcc, avaliador_id=avaliador_id)
        except AvaliacaoFase1.DoesNotExist:
            return Response({'detail': 'Avaliação não encontrada para este avaliador'}, status=status.HTTP_404_NOT_FOUND)

        # Filtrar apenas os campos de nota do request.data (excluir avaliador_id)
        dados = {k: v for k, v in request.data.items() if k != 'avaliador_id'}

        serializer = AvaliacaoFase1EscritaSerializer(avaliacao, data=dados, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        avaliacao_atualizada = serializer.save()

        return Response(AvaliacaoFase1Serializer(avaliacao_atualizada, context={'request': request}).data)

    @action(detail=True, methods=['post'], url_path='avaliacao-fase2/editar-coordenador', permission_classes=[IsCoordenador])
    def editar_avaliacao_fase2_coordenador(self, request, pk=None):
        """
        POST /api/tccs/{id}/avaliacao-fase2/editar-coordenador/
        Coordenador edita diretamente a avaliação de um avaliador na Fase II.
        Body: { avaliador_id, nota_coerencia_conteudo, nota_qualidade_apresentacao, nota_dominio_tema, nota_clareza_fluencia, nota_observancia_tempo, parecer, status }
        """
        from .models import AvaliacaoFase2
        from .serializers import AvaliacaoFase2Serializer, AvaliacaoFase2EscritaSerializer

        tcc = self.get_object()
        avaliador_id = request.data.get('avaliador_id')

        if not avaliador_id:
            return Response({'detail': 'avaliador_id é obrigatório'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            avaliacao = AvaliacaoFase2.objects.get(tcc=tcc, avaliador_id=avaliador_id)
        except AvaliacaoFase2.DoesNotExist:
            return Response({'detail': 'Avaliação não encontrada para este avaliador'}, status=status.HTTP_404_NOT_FOUND)

        dados = {k: v for k, v in request.data.items() if k != 'avaliador_id'}

        serializer = AvaliacaoFase2EscritaSerializer(avaliacao, data=dados, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        avaliacao_atualizada = serializer.save()

        return Response(AvaliacaoFase2Serializer(avaliacao_atualizada, context={'request': request}).data)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated, IsCoordenador])
    def exportar_dados(self, request):
        """
        GET /api/tccs/exportar_dados/?dados=true&monografia=true&documentos=true
        Exporta dados de todos os alunos com TCC em formato ZIP.

        Query params (todos true por padrão):
        - dados: inclui avaliacoes.txt (fichas de avaliação)
        - monografia: inclui última monografia
        - documentos: inclui documentos gerais (plano, termos, apresentação, relatório, etc.)

        Apenas coordenador tem acesso.
        """
        # Ler filtros dos query params
        incluir_dados = request.query_params.get('dados', 'true').lower() == 'true'
        incluir_monografia = request.query_params.get('monografia', 'true').lower() == 'true'
        incluir_documentos = request.query_params.get('documentos', 'true').lower() == 'true'

        # Tipos de documentos gerais (tudo exceto MONOGRAFIA e MONOGRAFIA_AVALIACAO)
        tipos_documentos_gerais = [
            TipoDocumento.PLANO_DESENVOLVIMENTO,
            TipoDocumento.TERMO_ACEITE,
            TipoDocumento.TERMO_SOLICITACAO_AVALIACAO,
            TipoDocumento.APRESENTACAO,
            TipoDocumento.ATA,
            TipoDocumento.RELATORIO_AVALIACAO,
            TipoDocumento.OUTRO,
        ]

        # Buscar todos os TCCs agrupados por aluno
        tccs = TCC.objects.select_related('aluno').order_by('aluno__nome_completo', '-criado_em')

        # Agrupar TCCs por aluno (pegar apenas o mais recente de cada aluno)
        alunos_tccs = {}
        for tcc in tccs:
            if tcc.aluno_id not in alunos_tccs:
                alunos_tccs[tcc.aluno_id] = tcc

        # Criar ZIP em memória
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Controlar nomes duplicados
            nomes_usados = defaultdict(int)

            for aluno_id, tcc in alunos_tccs.items():
                aluno = tcc.aluno
                nome_aluno = aluno.nome_completo

                # Gerar nome único da pasta (adicionar (2), (3) se duplicado)
                if nomes_usados[nome_aluno] == 0:
                    nome_pasta = nome_aluno
                else:
                    nome_pasta = f"{nome_aluno} ({nomes_usados[nome_aluno] + 1})"
                nomes_usados[nome_aluno] += 1

                # Monografia aprovada pelo orientador
                if incluir_monografia:
                    ultima_monografia = DocumentoTCC.objects.filter(
                        tcc=tcc,
                        tipo_documento=TipoDocumento.MONOGRAFIA,
                        status=StatusDocumento.APROVADO
                    ).order_by('-versao', '-criado_em').first()

                    if ultima_monografia and ultima_monografia.arquivo:
                        try:
                            arquivo_path = ultima_monografia.arquivo.path
                            arquivo_nome = ultima_monografia.nome_original or os.path.basename(arquivo_path)
                            with open(arquivo_path, 'rb') as f:
                                zip_file.writestr(f"{nome_pasta}/{arquivo_nome}", f.read())
                        except Exception:
                            pass

                # Dados (relatório txt)
                if incluir_dados:
                    avaliacoes_f1 = AvaliacaoFase1.objects.filter(tcc=tcc).select_related('avaliador')
                    avaliacoes_f2 = AvaliacaoFase2.objects.filter(tcc=tcc).select_related('avaliador')

                    conteudo_completo = self._gerar_texto_avaliacoes_completo(
                        avaliacoes_f1, avaliacoes_f2, aluno, tcc
                    )
                    zip_file.writestr(f"{nome_pasta}/dados.txt", conteudo_completo)

                # Documentos gerais
                if incluir_documentos:
                    documentos = DocumentoTCC.objects.filter(
                        tcc=tcc,
                        tipo_documento__in=tipos_documentos_gerais
                    ).order_by('tipo_documento', '-versao')

                    for doc in documentos:
                        if doc.arquivo:
                            try:
                                arquivo_path = doc.arquivo.path
                                arquivo_nome = doc.nome_original or os.path.basename(arquivo_path)
                                with open(arquivo_path, 'rb') as f:
                                    zip_file.writestr(f"{nome_pasta}/{arquivo_nome}", f.read())
                            except Exception:
                                pass

        # Preparar resposta com streaming
        zip_buffer.seek(0)
        response = StreamingHttpResponse(
            zip_buffer,
            content_type='application/zip'
        )
        response['Content-Disposition'] = 'attachment; filename="dados_tccs.zip"'

        return response

    @action(detail=True, methods=['get'], url_path='exportar-dados', permission_classes=[IsAuthenticated, IsCoordenador])
    def exportar_dados_tcc(self, request, pk=None):
        """
        GET /api/tccs/{id}/exportar-dados/?dados=true&monografia=true&documentos=true
        Exporta dados de um TCC individual em formato ZIP.

        Query params (todos true por padrão):
        - dados: inclui dados.txt (relatório completo)
        - monografia: inclui monografia aprovada
        - documentos: inclui documentos gerais (plano, termos, apresentação, relatório, etc.)
        """
        tcc = self.get_object()
        aluno = tcc.aluno
        nome_aluno = aluno.nome_completo

        # Ler filtros dos query params
        incluir_dados = request.query_params.get('dados', 'true').lower() == 'true'
        incluir_monografia = request.query_params.get('monografia', 'true').lower() == 'true'
        incluir_documentos = request.query_params.get('documentos', 'true').lower() == 'true'

        # Tipos de documentos gerais (tudo exceto MONOGRAFIA e MONOGRAFIA_AVALIACAO)
        tipos_documentos_gerais = [
            TipoDocumento.PLANO_DESENVOLVIMENTO,
            TipoDocumento.TERMO_ACEITE,
            TipoDocumento.TERMO_SOLICITACAO_AVALIACAO,
            TipoDocumento.APRESENTACAO,
            TipoDocumento.ATA,
            TipoDocumento.RELATORIO_AVALIACAO,
            TipoDocumento.OUTRO,
        ]

        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Monografia aprovada
            if incluir_monografia:
                ultima_monografia = DocumentoTCC.objects.filter(
                    tcc=tcc,
                    tipo_documento=TipoDocumento.MONOGRAFIA,
                    status=StatusDocumento.APROVADO
                ).order_by('-versao', '-criado_em').first()

                if ultima_monografia and ultima_monografia.arquivo:
                    try:
                        arquivo_path = ultima_monografia.arquivo.path
                        arquivo_nome = ultima_monografia.nome_original or os.path.basename(arquivo_path)
                        with open(arquivo_path, 'rb') as f:
                            zip_file.writestr(arquivo_nome, f.read())
                    except Exception:
                        pass

            # Relatório TXT
            if incluir_dados:
                avaliacoes_f1 = AvaliacaoFase1.objects.filter(tcc=tcc).select_related('avaliador')
                avaliacoes_f2 = AvaliacaoFase2.objects.filter(tcc=tcc).select_related('avaliador')

                conteudo = self._gerar_texto_avaliacoes_completo(
                    avaliacoes_f1, avaliacoes_f2, aluno, tcc
                )
                zip_file.writestr('dados.txt', conteudo)

            # Documentos gerais
            if incluir_documentos:
                documentos = DocumentoTCC.objects.filter(
                    tcc=tcc,
                    tipo_documento__in=tipos_documentos_gerais
                ).order_by('tipo_documento', '-versao')

                for doc in documentos:
                    if doc.arquivo:
                        try:
                            arquivo_path = doc.arquivo.path
                            arquivo_nome = doc.nome_original or os.path.basename(arquivo_path)
                            with open(arquivo_path, 'rb') as f:
                                zip_file.writestr(arquivo_nome, f.read())
                        except Exception:
                            pass

        zip_buffer.seek(0)
        nome_arquivo = f"{nome_aluno}.zip".replace(' ', '_')
        response = StreamingHttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        return response

    def _fmt(self, nota):
        """Formata uma nota de forma segura, retornando '-' se for None."""
        if nota is None:
            return "-"
        return f"{nota:.2f}".replace('.', ',')

    def _linha_nota(self, criterio, nota, peso_max, largura=45):
        """Gera linha formatada de nota com pontos alinhados."""
        nota_str = f"{self._fmt(nota)} / {self._fmt(peso_max)}"
        pontos = '.' * max(2, largura - len(criterio) - len(nota_str))
        return f"    {criterio} {pontos} {nota_str}"

    def _centralizar(self, texto, largura=78):
        """Centraliza texto dentro da largura, com prefixo de 2 espaços."""
        return "  " + texto.center(largura)

    def _gerar_texto_avaliacoes_completo(self, avaliacoes_f1, avaliacoes_f2, aluno, tcc):
        """Gera relatório completo do TCC em texto."""
        from definicoes.models import CalendarioSemestre
        from .models import AgendamentoDefesa

        cal = CalendarioSemestre.obter_calendario_atual(tcc.semestre)
        # Pesos Fase I
        p_res = float(cal.peso_resumo) if cal else 1.0
        p_int = float(cal.peso_introducao) if cal else 2.0
        p_rev = float(cal.peso_revisao) if cal else 2.0
        p_dev = float(cal.peso_desenvolvimento) if cal else 3.5
        p_con = float(cal.peso_conclusoes) if cal else 1.5
        # Pesos Fase II
        p_coe = float(cal.peso_coerencia_conteudo) if cal else 2.0
        p_qua = float(cal.peso_qualidade_apresentacao) if cal else 2.0
        p_dom = float(cal.peso_dominio_tema) if cal else 2.5
        p_cla = float(cal.peso_clareza_fluencia) if cal else 2.5
        p_tem = float(cal.peso_observancia_tempo) if cal else 1.0

        L = []
        sep = "  " + "\u2500" * 78  # ──────
        sep2 = "  " + "\u2550" * 78  # ══════

        # ── CABEÇALHO ──
        L.append(sep2)
        L.append(self._centralizar("RELATÓRIO DO TCC \u2014 TRABALHO DE CONCLUSÃO DE CURSO"))
        L.append(sep2)
        L.append("")
        L.append(f"    Aluno ............. {aluno.nome_completo}")
        orientador_nome = tcc.orientador.nome_completo if tcc.orientador else ''
        L.append(f"    Orientador ........ {orientador_nome}")

        # Co-orientador (só se tiver)
        coori_nome = ''
        if tcc.coorientador:
            coori_nome = tcc.coorientador.nome_completo
        elif tcc.coorientador_nome:
            coori_nome = tcc.coorientador_nome
        if coori_nome:
            L.append(f"    Co-orientador ..... {coori_nome}")

        L.append(f"    Tema .............. {tcc.titulo}")
        L.append(f"    Semestre .......... {tcc.semestre}")
        L.append("")

        # ── INFORMAÇÕES DO PROCESSO ──
        # Monografia
        monografia_aprovada = DocumentoTCC.objects.filter(
            tcc=tcc, tipo_documento=TipoDocumento.MONOGRAFIA, status=StatusDocumento.APROVADO
        ).exists()
        monografia_pendente = DocumentoTCC.objects.filter(
            tcc=tcc, tipo_documento=TipoDocumento.MONOGRAFIA
        ).exclude(status=StatusDocumento.APROVADO).exists()

        if monografia_aprovada:
            status_mono = 'Aprovada'
        elif monografia_pendente:
            status_mono = 'Aguardando ajustes'
        else:
            status_mono = ''
        L.append(f"    Monografia ................. {status_mono}")

        # Continuidade
        if tcc.etapa_atual == EtapaTCC.DESCONTINUADO:
            status_cont = 'Encerrada'
        elif tcc.flag_continuidade:
            status_cont = 'Ok'
        else:
            status_cont = ''
        L.append(f"    Confirmação de continuidade  {status_cont}")

        # Termo de solicitação
        tem_termo = DocumentoTCC.objects.filter(
            tcc=tcc, tipo_documento=TipoDocumento.TERMO_SOLICITACAO_AVALIACAO
        ).exists()
        L.append(f"    Termo de solicitação ........ {'Enviado' if tem_termo else ''}")

        # Fase I resultado
        fase1_resultado = ''
        if tcc.nf1 is not None:
            fase1_resultado = 'Aprovado' if float(tcc.nf1) >= 6 else 'Reprovado'
        elif tcc.etapa_atual == EtapaTCC.REPROVADO_FASE_1:
            fase1_resultado = 'Reprovado'
        L.append(f"    Fase I ..................... {fase1_resultado}")

        # Agendamento
        agendamento_txt = ''
        try:
            agendamento = tcc.agendamento_defesa
            data_str = agendamento.data.strftime('%d/%m/%Y')
            hora_str = agendamento.hora.strftime('%H:%M')
            agendamento_txt = f"{data_str} às {hora_str} - {agendamento.local}"
        except AgendamentoDefesa.DoesNotExist:
            pass
        L.append(f"    Agendamento ................ {agendamento_txt}")

        # Fase II resultado
        fase2_resultado = ''
        if tcc.nf2 is not None:
            fase2_resultado = 'Aprovado' if float(tcc.nf2) >= 6 else 'Reprovado'
        elif tcc.etapa_atual == EtapaTCC.REPROVADO_FASE_2:
            fase2_resultado = 'Reprovado'
        L.append(f"    Fase II .................... {fase2_resultado}")

        # Etapa atual
        L.append(f"    Etapa atual ................ {tcc.get_etapa_atual_display()}")
        L.append("")

        # ── FASE I ──
        if avaliacoes_f1.exists():
            L.append(sep)
            L.append(self._centralizar("FASE I \u2014 AVALIAÇÃO DA MONOGRAFIA"))
            L.append(sep)
            L.append("")

            notas_totais_f1 = []
            for i, av in enumerate(avaliacoes_f1, 1):
                trat = av.avaliador.tratamento or ''
                nome = f"{trat} {av.avaliador.nome_completo}".strip()
                L.append(f"    Avaliador #{i}: {nome}")
                L.append(f"    {'─' * (len(nome) + 16)}")

                if av.status in ['ENVIADO', 'BLOQUEADO']:
                    L.append(self._linha_nota("Resumo", av.nota_resumo, p_res))
                    L.append(self._linha_nota("Introdução/Relevância", av.nota_introducao, p_int))
                    L.append(self._linha_nota("Revisão Bibliográfica", av.nota_revisao, p_rev))
                    L.append(self._linha_nota("Desenvolvimento", av.nota_desenvolvimento, p_dev))
                    L.append(self._linha_nota("Conclusões", av.nota_conclusoes, p_con))
                    nota_total = av.calcular_nota_total()
                    notas_totais_f1.append(float(nota_total) if nota_total else 0)
                    total_txt = f"TOTAL {self._fmt(nota_total)} / 10,00"
                    total_pad = ' ' * (53 - len(total_txt))
                    L.append(f"{total_pad}{'─' * len(total_txt)}")
                    L.append(f"{total_pad}{total_txt}")

                    # Comentário geral do avaliador
                    if av.parecer and av.parecer.strip():
                        L.append("")
                        for linha in av.parecer.strip().split('\n'):
                            L.append(f"      {linha}")
                else:
                    L.append("      (Avaliação pendente)")

                L.append("")

            if notas_totais_f1:
                nf1 = sum(notas_totais_f1) / len(notas_totais_f1)
                nf1_pond = nf1 * 0.6
                L.append(f"    \u25b8 NF1 (Média Fase I) = {self._fmt(nf1)}")
                L.append(f"    \u25b8 NF1 (Nota final I) = {self._fmt(nf1_pond)}")
                L.append("")

        # ── FASE II ──
        if avaliacoes_f2.exists():
            L.append(sep)
            L.append(self._centralizar("FASE II \u2014 AVALIAÇÃO DA APRESENTAÇÃO"))
            L.append(sep)
            L.append("")

            notas_totais_f2 = []
            for i, av in enumerate(avaliacoes_f2, 1):
                trat = av.avaliador.tratamento or ''
                nome = f"{trat} {av.avaliador.nome_completo}".strip()
                L.append(f"    Avaliador #{i}: {nome}")
                L.append(f"    {'─' * (len(nome) + 16)}")

                if av.status in ['ENVIADO', 'BLOQUEADO']:
                    L.append(self._linha_nota("Coerência do Conteúdo", av.nota_coerencia_conteudo, p_coe))
                    L.append(self._linha_nota("Qualidade e Estrutura", av.nota_qualidade_apresentacao, p_qua))
                    L.append(self._linha_nota("Domínio e Conhecimento", av.nota_dominio_tema, p_dom))
                    L.append(self._linha_nota("Clareza e Fluência", av.nota_clareza_fluencia, p_cla))
                    L.append(self._linha_nota("Observância do Tempo", av.nota_observancia_tempo, p_tem))
                    nota_total = av.calcular_nota_total()
                    notas_totais_f2.append(float(nota_total) if nota_total else 0)
                    total_txt = f"TOTAL {self._fmt(nota_total)} / 10,00"
                    total_pad = ' ' * (53 - len(total_txt))
                    L.append(f"{total_pad}{'─' * len(total_txt)}")
                    L.append(f"{total_pad}{total_txt}")

                    # Comentário geral do avaliador
                    if av.parecer and av.parecer.strip():
                        L.append("")
                        for linha in av.parecer.strip().split('\n'):
                            L.append(f"      {linha}")
                else:
                    L.append("      (Avaliação pendente)")

                L.append("")

            if notas_totais_f2:
                media_f2 = sum(notas_totais_f2) / len(notas_totais_f2)
                nf2_pond = media_f2 * 0.4
                L.append(f"    \u25b8 NF2 (Média Fase II) = {self._fmt(media_f2)}")
                L.append(f"    \u25b8 NF2 (Nota final II) = {self._fmt(nf2_pond)}")
                L.append("")

        # ── RESULTADO FINAL ──
        if avaliacoes_f1.exists() and avaliacoes_f2.exists():
            nf1_val = sum(notas_totais_f1) / len(notas_totais_f1) if notas_totais_f1 else 0
            nf2_val = sum(notas_totais_f2) / len(notas_totais_f2) if notas_totais_f2 else 0
            nf1_pond = nf1_val * 0.6
            nf2_pond = nf2_val * 0.4
            nota_final = nf1_pond + nf2_pond
            resultado = tcc.resultado_final or ('APROVADO' if nota_final >= 6 else 'REPROVADO')

            L.append(sep)
            L.append(self._centralizar("RESULTADO FINAL"))
            L.append(sep)
            L.append("")
            L.append(f"      NF1 (Fase I)  = {self._fmt(nf1_val)}  \u00d7  0,6  =  {self._fmt(nf1_pond)}")
            L.append(f"      NF2 (Fase II) = {self._fmt(nf2_val)}  \u00d7  0,4  =  {self._fmt(nf2_pond)}")
            nf_txt = f"NOTA FINAL {self._fmt(nota_final)}"
            nf_pad = ' ' * (53 - len(nf_txt))
            L.append(f"{nf_pad}{'─' * len(nf_txt)}")
            L.append(f"{nf_pad}{nf_txt}")
            L.append("")
            L.append(f"      Resultado: {resultado}")
            L.append("")

        L.append(sep2)
        return "\n".join(L)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated, IsCoordenador])
    def verificar_email_reset(self, request):
        """
        GET /api/tccs/verificar_email_reset/
        Verifica se há email configurado para envio do backup.
        """
        from definicoes.models import ConfiguracaoEmail
        config = ConfiguracaoEmail.obter_configuracao()
        email_configurado = bool(config.email_host_user)
        return Response({
            'email_configurado': email_configurado,
            'email': config.email_host_user if email_configurado else None
        })

    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated, IsCoordenador])
    def resetar_periodo(self, request):
        """
        POST /api/tccs/resetar_periodo/
        Reseta o período: faz backup dos dados, envia por email e apaga tudo
        exceto professores e coordenador.
        """
        logger = logging.getLogger(__name__)
        from users.models import Usuario
        from avisos.models import Aviso, ComentarioAviso
        from notificacoes.models import Notificacao, PreferenciasEmail
        from definicoes.models import ConfiguracaoEmail

        senha = request.data.get('senha')
        email_destino = request.data.get('email_destino')

        if not senha:
            return Response({'detail': 'Senha é obrigatória.'}, status=status.HTTP_400_BAD_REQUEST)

        # Validar senha do coordenador
        if not check_password(senha, request.user.password):
            return Response({'detail': 'Senha incorreta.'}, status=status.HTTP_403_FORBIDDEN)

        # Verificar email de destino
        config_email = ConfiguracaoEmail.obter_configuracao()
        if not email_destino:
            if config_email.email_host_user:
                email_destino = config_email.email_host_user
            else:
                return Response(
                    {'detail': 'Nenhum email configurado. Envie email_destino no body.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # ============ GERAR BACKUP ZIP (mesmo que exportar_dados) ============
        tccs_qs = TCC.objects.select_related('aluno', 'orientador', 'coorientador').order_by('aluno__nome_completo', '-criado_em')
        alunos_tccs = {}
        for tcc in tccs_qs:
            if tcc.aluno_id not in alunos_tccs:
                alunos_tccs[tcc.aluno_id] = tcc

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            nomes_usados = defaultdict(int)

            for aluno_id, tcc in alunos_tccs.items():
                aluno = tcc.aluno
                nome_aluno = aluno.nome_completo
                if nomes_usados[nome_aluno] == 0:
                    nome_pasta = nome_aluno
                else:
                    nome_pasta = f"{nome_aluno} ({nomes_usados[nome_aluno] + 1})"
                nomes_usados[nome_aluno] += 1

                # Monografia
                ultima_monografia = DocumentoTCC.objects.filter(
                    tcc=tcc, tipo_documento=TipoDocumento.MONOGRAFIA
                ).order_by('-versao', '-criado_em').first()

                if ultima_monografia and ultima_monografia.arquivo:
                    try:
                        arquivo_path = ultima_monografia.arquivo.path
                        arquivo_nome = ultima_monografia.nome_original or os.path.basename(arquivo_path)
                        with open(arquivo_path, 'rb') as f:
                            zip_file.writestr(f"{nome_pasta}/{arquivo_nome}", f.read())
                    except Exception:
                        pass

                # Relatório TXT
                avaliacoes_f1 = AvaliacaoFase1.objects.filter(tcc=tcc).select_related('avaliador')
                avaliacoes_f2 = AvaliacaoFase2.objects.filter(tcc=tcc).select_related('avaliador')
                conteudo = self._gerar_texto_avaliacoes_completo(avaliacoes_f1, avaliacoes_f2, aluno, tcc)
                zip_file.writestr(f"{nome_pasta}/dados.txt", conteudo)

            # ============ GERAR EXCEL (Relatório) ============
            try:
                excel_buffer = self._gerar_excel_relatorio()
                zip_file.writestr('Relatorio_TCCs.xlsx', excel_buffer.getvalue())
            except Exception as e:
                logger.error(f"Erro ao gerar Excel: {e}")

        zip_buffer.seek(0)
        zip_bytes = zip_buffer.getvalue()

        # ============ GERAR TXT CONSOLIDADO ============
        txt_consolidado = self._gerar_txt_consolidado(alunos_tccs)

        # ============ ENVIAR EMAIL ============
        try:
            # Forçar backend SMTP real, mesmo que email_enabled esteja False
            from notificacoes.email_backend import EmailBackend as SMTPBackend

            # Configurar credenciais SMTP a partir da configuração salva
            smtp_settings = {
                'host': config_email.email_host,
                'port': config_email.email_port,
                'username': config_email.email_host_user,
                'password': config_email.get_password(),
                'use_tls': config_email.email_use_tls,
            }
            from_email = f'Portal TCC <{config_email.email_host_user}>'

            email = DjangoEmailMessage(
                subject='Backup do Período - Portal TCC',
                body='Segue em anexo o backup dos dados do período que foi resetado.\n\n'
                     '- avaliacoes_consolidado.txt: Dados de todos os alunos em texto\n'
                     '- Relatorio_TCCs.xlsx: Relatório completo em Excel\n\n'
                     'Este email foi gerado automaticamente pelo Portal TCC.',
                from_email=from_email,
                to=[email_destino],
            )
            email.attach('avaliacoes_consolidado.txt', txt_consolidado.encode('utf-8'), 'text/plain; charset=utf-8')

            try:
                excel_buffer_email = self._gerar_excel_relatorio()
                email.attach(
                    'Relatorio_TCCs.xlsx',
                    excel_buffer_email.getvalue(),
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                logger.error(f"Erro ao gerar Excel para email: {e}")

            # Enviar usando backend SMTP direto (ignora email_enabled)
            backend = SMTPBackend(
                host=smtp_settings['host'],
                port=smtp_settings['port'],
                username=smtp_settings['username'],
                password=smtp_settings['password'],
                use_tls=smtp_settings['use_tls'],
                fail_silently=False,
            )
            backend.send_messages([email])
            email_enviado = True
        except Exception as e:
            logger.error(f"Erro ao enviar email de backup: {e}")
            email_enviado = False

        # ============ APAGAR DADOS ============
        # Contar antes de apagar
        contagem = {
            'tccs': TCC.objects.count(),
            'alunos': Usuario.objects.filter(tipo_usuario='ALUNO').count(),
            'externos': Usuario.objects.filter(tipo_usuario='AVALIADOR').count(),
            'avisos': Aviso.objects.count(),
            'notificacoes': Notificacao.objects.count(),
        }

        # Apagar TCCs (cascade deleta: DocumentoTCC, EventoTimeline, BancaFase1,
        # MembroBanca, AvaliacaoFase1, AvaliacaoFase2, AgendamentoDefesa, HistoricoRecusa, etc.)
        TCC.objects.all().delete()

        # Apagar solicitações de orientação restantes
        SolicitacaoOrientacao.objects.all().delete()

        # Apagar avisos e comentários
        Aviso.objects.all().delete()

        # Apagar notificações
        Notificacao.objects.all().delete()

        # Apagar alunos e avaliadores externos
        Usuario.objects.filter(tipo_usuario='ALUNO').delete()
        Usuario.objects.filter(tipo_usuario='AVALIADOR').delete()

        # Apagar arquivos físicos
        media_tccs_path = os.path.join(settings.MEDIA_ROOT, 'tccs')
        if os.path.exists(media_tccs_path):
            shutil.rmtree(media_tccs_path)
            os.makedirs(media_tccs_path)

        # Retornar ZIP para download
        response = StreamingHttpResponse(io.BytesIO(zip_bytes), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="backup_periodo.zip"'
        response['X-Email-Enviado'] = 'true' if email_enviado else 'false'
        response['X-Contagem'] = str(contagem)
        response['Access-Control-Expose-Headers'] = 'X-Email-Enviado, X-Contagem'
        return response

    def _gerar_txt_consolidado(self, alunos_tccs):
        """Gera um TXT consolidado com dados de todos os alunos."""
        partes = []
        for aluno_id, tcc in alunos_tccs.items():
            aluno = tcc.aluno
            avaliacoes_f1 = AvaliacaoFase1.objects.filter(tcc=tcc).select_related('avaliador')
            avaliacoes_f2 = AvaliacaoFase2.objects.filter(tcc=tcc).select_related('avaliador')
            conteudo = self._gerar_texto_avaliacoes_completo(avaliacoes_f1, avaliacoes_f2, aluno, tcc)
            partes.append(conteudo)
        return "\n\n".join(partes)

    def _gerar_excel_relatorio(self):
        """Gera Excel com as mesmas 5 abas do Relatórios do frontend."""
        import openpyxl
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # Buscar dados
        tccs = TCC.objects.select_related(
            'aluno', 'orientador', 'coorientador'
        ).order_by('id')

        bancas = {}
        for banca in BancaFase1.objects.select_related('tcc').prefetch_related(
            'membros__usuario'
        ):
            bancas[banca.tcc_id] = banca

        avaliacoes_f1_map = {}
        for av in AvaliacaoFase1.objects.select_related('avaliador'):
            avaliacoes_f1_map.setdefault(av.tcc_id, []).append(av)

        avaliacoes_f2_map = {}
        for av in AvaliacaoFase2.objects.select_related('avaliador'):
            avaliacoes_f2_map.setdefault(av.tcc_id, []).append(av)

        def nota_num(valor):
            if valor is None:
                return None
            try:
                n = float(valor)
                return round(n, 2)
            except (TypeError, ValueError):
                return None

        # === ABA 1: Dados Gerais ===
        ws1 = wb.active
        ws1.title = 'Dados Gerais'
        headers1 = ['ID', 'Curso', 'Título', 'Aluno', 'Orientador', 'Orientador - Tratamento',
                     'Orientador - Afiliação', 'Coorientador', 'Coorientador - Tratamento',
                     'Coorientador - Afiliação', 'Avaliador 1 (Fase I)', 'Avaliador 2 (Fase I)',
                     'Avaliador 1 (Fase II)', 'Avaliador 2 (Fase II)', 'Avaliador 3 (Fase II)',
                     'Data Defesa', 'Semestre']
        ws1.append(headers1)

        for tcc in tccs:
            banca = bancas.get(tcc.id)
            avaliadores = sorted(
                [m for m in (banca.membros.all() if banca else []) if m.tipo == 'AVALIADOR'],
                key=lambda m: m.ordem
            )
            coori = tcc.coorientador
            ws1.append([
                tcc.id,
                tcc.aluno.get_curso_display() if tcc.aluno.curso else '',
                tcc.titulo,
                tcc.aluno.nome_completo,
                tcc.orientador.nome_completo if tcc.orientador else '',
                tcc.orientador.tratamento if tcc.orientador else '',
                getattr(tcc.orientador, 'afiliacao', '') if tcc.orientador else '',
                coori.nome_completo if coori else '',
                coori.tratamento if coori else '',
                getattr(coori, 'afiliacao', '') if coori else '',
                avaliadores[0].usuario.nome_completo if len(avaliadores) > 0 else '',
                avaliadores[1].usuario.nome_completo if len(avaliadores) > 1 else '',
                tcc.orientador.nome_completo if tcc.orientador else '',
                avaliadores[0].usuario.nome_completo if len(avaliadores) > 0 else '',
                avaliadores[1].usuario.nome_completo if len(avaliadores) > 1 else '',
                tcc.data_defesa.strftime('%d/%m/%Y') if tcc.data_defesa else '',
                tcc.semestre or '',
            ])

        # === ABA 2: Avaliações Fase I ===
        ws2 = wb.create_sheet('Avaliações Fase I')
        headers2 = ['ID', 'Título', 'Aluno', 'Orientador', 'Avaliador',
                     'Q1 - Resumo', 'Q2 - Introdução', 'Q3 - Revisão',
                     'Q4 - Desenvolvimento', 'Q5 - Conclusões', 'Nota Total']
        ws2.append(headers2)

        for tcc in tccs:
            banca = bancas.get(tcc.id)
            avaliadores = sorted(
                [m for m in (banca.membros.all() if banca else []) if m.tipo == 'AVALIADOR'],
                key=lambda m: m.ordem
            )
            avaliacoes = avaliacoes_f1_map.get(tcc.id, [])

            if not avaliadores:
                ws2.append([tcc.id, tcc.titulo, tcc.aluno.nome_completo,
                           tcc.orientador.nome_completo if tcc.orientador else '',
                           '', None, None, None, None, None, None])
            else:
                for membro in avaliadores:
                    aval = next((a for a in avaliacoes if a.avaliador_id == membro.usuario_id), None)
                    ws2.append([
                        tcc.id, tcc.titulo, tcc.aluno.nome_completo,
                        tcc.orientador.nome_completo if tcc.orientador else '',
                        membro.usuario.nome_completo,
                        nota_num(aval.nota_resumo) if aval else None,
                        nota_num(aval.nota_introducao) if aval else None,
                        nota_num(aval.nota_revisao) if aval else None,
                        nota_num(aval.nota_desenvolvimento) if aval else None,
                        nota_num(aval.nota_conclusoes) if aval else None,
                        nota_num(aval.calcular_nota_total()) if aval and aval.status in ['ENVIADO', 'BLOQUEADO'] else None,
                    ])

        # === ABA 3: Avaliações Fase II ===
        ws3 = wb.create_sheet('Avaliações Fase II')
        headers3 = ['ID', 'Título', 'Aluno', 'Orientador', 'Avaliador',
                     'Q1 - Coerência', 'Q2 - Qualidade', 'Q3 - Domínio',
                     'Q4 - Clareza', 'Q5 - Tempo', 'Nota Total']
        ws3.append(headers3)

        for tcc in tccs:
            banca = bancas.get(tcc.id)
            avaliadores_f1 = sorted(
                [m for m in (banca.membros.all() if banca else []) if m.tipo == 'AVALIADOR'],
                key=lambda m: m.ordem
            )
            avaliacoes = avaliacoes_f2_map.get(tcc.id, [])

            # Orientador + 2 avaliadores da fase 1
            aval_orient = next((a for a in avaliacoes if a.avaliador_id == (tcc.orientador_id or -1)), None)
            if not aval_orient and tcc.coorientador_id:
                aval_orient = next((a for a in avaliacoes if a.avaliador_id == tcc.coorientador_id), None)

            lista_avals = [
                (tcc.orientador.nome_completo if tcc.orientador else '', aval_orient),
            ]
            for membro in avaliadores_f1:
                aval = next((a for a in avaliacoes if a.avaliador_id == membro.usuario_id), None)
                lista_avals.append((membro.usuario.nome_completo, aval))

            for nome, aval in lista_avals:
                ws3.append([
                    tcc.id, tcc.titulo, tcc.aluno.nome_completo,
                    tcc.orientador.nome_completo if tcc.orientador else '',
                    nome,
                    nota_num(aval.nota_coerencia_conteudo) if aval else None,
                    nota_num(aval.nota_qualidade_apresentacao) if aval else None,
                    nota_num(aval.nota_dominio_tema) if aval else None,
                    nota_num(aval.nota_clareza_fluencia) if aval else None,
                    nota_num(aval.nota_observancia_tempo) if aval else None,
                    nota_num(aval.calcular_nota_total()) if aval and aval.status in ['ENVIADO', 'BLOQUEADO'] else None,
                ])

        # === ABA 4: Apuração Final ===
        ws4 = wb.create_sheet('Apuração Final')
        headers4 = ['ID', 'Título', 'Aluno', 'Orientador',
                     'Fase I - Nota 1', 'Fase I - Nota 2', 'Fase I - Média', 'Fase I - Nota com peso (60%)',
                     'Fase II - Nota 1', 'Fase II - Nota 2', 'Fase II - Nota 3',
                     'Fase II - Média', 'Fase II - Nota com peso (40%)', 'Nota Final']
        ws4.append(headers4)

        for tcc in tccs:
            avals_f1 = avaliacoes_f1_map.get(tcc.id, [])
            notas_f1 = [nota_num(a.calcular_nota_total()) for a in avals_f1 if a.status in ['ENVIADO', 'BLOQUEADO']]
            n1f1 = notas_f1[0] if len(notas_f1) > 0 else None
            n2f1 = notas_f1[1] if len(notas_f1) > 1 else None
            media_f1 = round((n1f1 + n2f1) / 2, 2) if n1f1 is not None and n2f1 is not None else None
            peso_f1 = round(media_f1 * 0.6, 2) if media_f1 is not None else None

            avals_f2 = avaliacoes_f2_map.get(tcc.id, [])
            nota_orient = None
            for a in avals_f2:
                if a.avaliador_id == (tcc.orientador_id or -1) and a.status in ['ENVIADO', 'BLOQUEADO']:
                    nota_orient = nota_num(a.calcular_nota_total())
                    break
            if nota_orient is None and tcc.coorientador_id:
                for a in avals_f2:
                    if a.avaliador_id == tcc.coorientador_id and a.status in ['ENVIADO', 'BLOQUEADO']:
                        nota_orient = nota_num(a.calcular_nota_total())
                        break

            outros_f2 = [a for a in avals_f2 if a.avaliador_id != (tcc.orientador_id or -1) and a.status in ['ENVIADO', 'BLOQUEADO']]
            outros_f2.sort(key=lambda a: a.id)
            n1f2 = nota_num(outros_f2[0].calcular_nota_total()) if len(outros_f2) > 0 else None
            n2f2 = nota_num(outros_f2[1].calcular_nota_total()) if len(outros_f2) > 1 else None

            media_f2 = round((nota_orient + n1f2 + n2f2) / 3, 2) if all(v is not None for v in [nota_orient, n1f2, n2f2]) else None
            peso_f2 = round(media_f2 * 0.4, 2) if media_f2 is not None else None
            nota_final = round(peso_f1 + peso_f2, 2) if peso_f1 is not None and peso_f2 is not None else None

            ws4.append([
                tcc.id, tcc.titulo, tcc.aluno.nome_completo,
                tcc.orientador.nome_completo if tcc.orientador else '',
                n1f1, n2f1, media_f1, peso_f1,
                nota_orient, n1f2, n2f2, media_f2, peso_f2, nota_final,
            ])

        # === ABA 5: Relatório de Avaliação ===
        ws5 = wb.create_sheet('Relatório de Avaliação')
        headers5 = ['ID', 'Título', 'Aluno', 'Data Defesa',
                     'Avaliador 1 (Fase I)', 'Parecer Avaliador 1 (Fase I)',
                     'Avaliador 2 (Fase I)', 'Parecer Avaliador 2 (Fase I)',
                     'Avaliador 1 (Fase II)', 'Parecer Avaliador 1 (Fase II)',
                     'Avaliador 2 (Fase II)', 'Parecer Avaliador 2 (Fase II)',
                     'Avaliador 3 (Fase II)', 'Parecer Avaliador 3 (Fase II)']
        ws5.append(headers5)

        for tcc in tccs:
            banca = bancas.get(tcc.id)
            avaliadores_banca = sorted(
                [m for m in (banca.membros.all() if banca else []) if m.tipo == 'AVALIADOR'],
                key=lambda m: m.ordem
            )
            avals_f1 = avaliacoes_f1_map.get(tcc.id, [])
            avals_f2 = avaliacoes_f2_map.get(tcc.id, [])

            aval1_f1 = next((a for a in avals_f1 if len(avaliadores_banca) > 0 and a.avaliador_id == avaliadores_banca[0].usuario_id), None)
            aval2_f1 = next((a for a in avals_f1 if len(avaliadores_banca) > 1 and a.avaliador_id == avaliadores_banca[1].usuario_id), None)

            aval1_f2 = next((a for a in avals_f2 if a.avaliador_id == (tcc.orientador_id or -1)), None)
            if not aval1_f2 and tcc.coorientador_id:
                aval1_f2 = next((a for a in avals_f2 if a.avaliador_id == tcc.coorientador_id), None)
            aval2_f2 = next((a for a in avals_f2 if len(avaliadores_banca) > 0 and a.avaliador_id == avaliadores_banca[0].usuario_id), None)
            aval3_f2 = next((a for a in avals_f2 if len(avaliadores_banca) > 1 and a.avaliador_id == avaliadores_banca[1].usuario_id), None)

            ws5.append([
                tcc.id, tcc.titulo, tcc.aluno.nome_completo,
                tcc.data_defesa.strftime('%d/%m/%Y') if tcc.data_defesa else '',
                avaliadores_banca[0].usuario.nome_completo if len(avaliadores_banca) > 0 else '',
                aval1_f1.parecer if aval1_f1 and aval1_f1.parecer else '',
                avaliadores_banca[1].usuario.nome_completo if len(avaliadores_banca) > 1 else '',
                aval2_f1.parecer if aval2_f1 and aval2_f1.parecer else '',
                tcc.orientador.nome_completo if tcc.orientador else '',
                aval1_f2.parecer if aval1_f2 and aval1_f2.parecer else '',
                avaliadores_banca[0].usuario.nome_completo if len(avaliadores_banca) > 0 else '',
                aval2_f2.parecer if aval2_f2 and aval2_f2.parecer else '',
                avaliadores_banca[1].usuario.nome_completo if len(avaliadores_banca) > 1 else '',
                aval3_f2.parecer if aval3_f2 and aval3_f2.parecer else '',
            ])

        # Ajustar largura das colunas
        for ws in [ws1, ws2, ws3, ws4, ws5]:
            for col_idx, col in enumerate(ws.columns, 1):
                max_len = 10
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 60))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def reset(self, request):
        """
        POST /api/tccs/reset/
        DESENVOLVIMENTO APENAS: Remove todos os dados de TCC (TCCs, solicitações, documentos, eventos)
        e limpa arquivos de upload. Mantém usuários.
        """
        # Proteção: só permite em modo DEBUG
        if not settings.DEBUG:
            return Response(
                {'detail': 'Endpoint disponível apenas em modo de desenvolvimento'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Deletar todas as instâncias (cascade deleta relacionados)
        tcc_count = TCC.objects.count()
        solicitacao_count = SolicitacaoOrientacao.objects.count()
        documento_count = DocumentoTCC.objects.count()
        evento_count = EventoTimeline.objects.count()

        TCC.objects.all().delete()

        # Remover arquivos de media/tccs
        media_tccs_path = os.path.join(settings.MEDIA_ROOT, 'tccs')
        if os.path.exists(media_tccs_path):
            shutil.rmtree(media_tccs_path)
            os.makedirs(media_tccs_path)

        return Response({
            'message': 'Ambiente de TCC resetado com sucesso',
            'deletados': {
                'tccs': tcc_count,
                'solicitacoes': solicitacao_count,
                'documentos': documento_count,
                'eventos': evento_count
            }
        }, status=status.HTTP_200_OK)


class SolicitacaoOrientacaoViewSet(viewsets.ModelViewSet):
    """ViewSet para gerenciar Solicitacoes de Orientacao."""

    queryset = SolicitacaoOrientacao.objects.all()
    serializer_class = SolicitacaoOrientacaoSerializer
    permission_classes = [IsAuthenticated, IsSolicitacaoOwnerOrCoordenador]

    def perform_create(self, serializer):
        """Validar que apenas aluno dono do TCC pode criar solicitação."""
        tcc = serializer.validated_data.get('tcc')

        # Validar que usuário é aluno
        if self.request.user.tipo_usuario != 'ALUNO':
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Apenas alunos podem criar solicitações de orientação')

        # Validar que o TCC pertence ao aluno logado
        if tcc.aluno != self.request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Você não pode criar solicitação para TCC de outro aluno')

        solicitacao = serializer.save()

        # Notificar coordenador(es) sobre nova solicitação de orientação
        from users.models import Usuario
        from notificacoes.services import criar_notificacao_em_massa_com_email
        coordenadores = Usuario.objects.filter(tipo_usuario='COORDENADOR')
        criar_notificacao_em_massa_com_email(
            usuarios=list(coordenadores),
            tipo=TipoNotificacao.CONVITE_ALUNO,
            titulo='Nova Solicitação de Orientação',
            mensagem=f'O aluno {self.request.user.nome_completo} enviou uma solicitação de orientação para o professor {solicitacao.professor.nome_completo}.',
            campo_preferencia='coord_convite_aluno',
            action_url='/solicitacoes',
            tcc_id=tcc.id,
            prioridade=PrioridadeNotificacao.NORMAL
        )

    def get_queryset(self):
        """Filtrar solicitações conforme tipo de usuário."""
        usuario = self.request.user

        if usuario.tipo_usuario == 'COORDENADOR':
            return SolicitacaoOrientacao.objects.all()

        elif usuario.tipo_usuario == 'ALUNO':
            # Aluno vê solicitações dos seus TCCs
            return SolicitacaoOrientacao.objects.filter(tcc__aluno=usuario)

        elif usuario.tipo_usuario == 'PROFESSOR':
            # Professor vê solicitações recebidas
            return SolicitacaoOrientacao.objects.filter(professor=usuario)

        return SolicitacaoOrientacao.objects.none()

    @action(detail=False, methods=['get'], permission_classes=[IsProfessorOrCoordenador])
    def pendentes(self, request):
        """
        GET /api/solicitacoes/pendentes/
        Lista solicitacoes pendentes:
        - Coordenador: ve todas as solicitacoes pendentes e pode aceitar/recusar
        - Professor: ve suas solicitacoes pendentes (apenas informativo)
          A ação de aceitar/recusar continua com o coordenador.
        """
        usuario = request.user

        if usuario.tipo_usuario == 'COORDENADOR':
            # Coordenador ve todas as pendentes
            solicitacoes = SolicitacaoOrientacao.objects.filter(
                status=StatusSolicitacao.PENDENTE
            ).order_by('-criado_em')
        else:
            # Professor vê suas solicitações pendentes (apenas informativo).
            # Ação de aprovar/recusar continua com o coordenador.
            solicitacoes = SolicitacaoOrientacao.objects.filter(
                professor=usuario,
                status=StatusSolicitacao.PENDENTE
            ).order_by('-criado_em')

        serializer = self.get_serializer(solicitacoes, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], permission_classes=[IsCoordenador])
    def aceitar(self, request, pk=None):
        """
        POST /api/solicitacoes/{id}/aceitar/
        Coordenador aprova solicitacao de orientacao.
        """
        from django.db import transaction

        solicitacao = self.get_object()

        if solicitacao.status != StatusSolicitacao.PENDENTE:
            return Response(
                {'detail': 'Apenas solicitacoes pendentes podem ser aceitas'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Usar transação atômica para garantir consistência
        with transaction.atomic():
            # Atualizar solicitacao
            solicitacao.status = StatusSolicitacao.ACEITA
            solicitacao.respondido_em = timezone.now()
            solicitacao.resposta_professor = request.data.get('resposta', '')
            solicitacao.save()

            # Atualizar TCC
            tcc = solicitacao.tcc
            tcc.orientador = solicitacao.professor
            tcc.etapa_atual = EtapaTCC.DESENVOLVIMENTO

            # Resetar/garantir flags (inicio do desenvolvimento)
            tcc.flag_continuidade = False
            tcc.flag_liberado_avaliacao = False
            tcc.avaliacao_fase1_bloqueada = False

            # Se tem dados de coorientador na solicitacao, copiar para o TCC
            if solicitacao.coorientador_nome:
                tcc.coorientador_nome = solicitacao.coorientador_nome
                tcc.coorientador_titulacao = solicitacao.coorientador_titulacao
                tcc.coorientador_afiliacao = solicitacao.coorientador_afiliacao
                tcc.coorientador_lattes = solicitacao.coorientador_lattes

                # Tentar vincular coorientador cadastrado pelo nome
                from users.models import Usuario
                coorientador_user = Usuario.objects.filter(
                    nome_completo=solicitacao.coorientador_nome,
                    tipo_usuario__in=['PROFESSOR', 'COORDENADOR', 'AVALIADOR']
                ).first()
                if coorientador_user:
                    tcc.coorientador = coorientador_user

            tcc.save()

            # Auto-aprovar documentos iniciais (PLANO_DESENVOLVIMENTO e TERMO_ACEITE)
            DocumentoTCC.objects.filter(
                tcc=tcc,
                tipo_documento__in=['PLANO_DESENVOLVIMENTO', 'TERMO_ACEITE']
            ).update(
                status=StatusDocumento.APROVADO,
                feedback=''
            )

            # Criar evento
            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=request.user,
                tipo_evento=TipoEvento.SOLICITACAO_ACEITA,
                descricao=f'Solicitacao aprovada pelo coordenador {request.user.nome_completo}. Orientador atribuido: {solicitacao.professor.nome_completo}. Etapa: Desenvolvimento',
                detalhes_json={
                    'solicitacao_id': solicitacao.id,
                    'coordenador_id': request.user.id,
                    'orientador_id': solicitacao.professor.id
                },
                visibilidade=Visibilidade.TODOS
            )

            # Criar notificações e enviar e-mails
            from notificacoes.services import criar_notificacao_com_email
            from definicoes.models import CalendarioSemestre

            # Buscar datas do calendário (usa o calendário ativo mais recente)
            calendario = CalendarioSemestre.obter_calendario_atual()
            data_continuidade = calendario.avaliacao_continuidade_fim.strftime('%d/%m/%Y') if calendario and calendario.avaliacao_continuidade_fim else 'Não definida'
            data_submissao = calendario.submissao_monografia_fim.strftime('%d/%m/%Y') if calendario and calendario.submissao_monografia_fim else 'Não definida'
            obs_data = f' ({data_continuidade})' if calendario and calendario.avaliacao_continuidade_fim else ''

            # Dados do co-orientador (se existir)
            tem_coorientador = bool(tcc.coorientador_nome or tcc.coorientador)
            coorientador_nome = ''
            if tcc.coorientador:
                coorientador_nome = tcc.coorientador.nome_completo
            elif tcc.coorientador_nome:
                coorientador_nome = tcc.coorientador_nome

            # Linha de co-orientador (condicional)
            linha_coori_html = f'<br><b>Co-orientador:</b> {coorientador_nome}' if tem_coorientador else ''
            linha_coori_texto = f'\nCo-orientador: {coorientador_nome}' if tem_coorientador else ''

            # Bloco de prazos e observação (compartilhado)
            bloco_prazos_html = f"""<p><b>📅 Próximos prazos:</b></p>

<p>{data_continuidade} – Confirmação de continuidade<br>
{data_submissao} – Último dia para submissão da <b>Monografia aprovada pelo orientador</b> e do <b>Termo de Solicitação de Avaliação da Monografia</b></p>

<p><b>⚠ Observação importante:</b></p>

<p>Caso você envie o seu TCC para análise e ele seja aprovado pelo orientador antes da data prevista para confirmação de continuidade, a funcionalidade terá sua abertura antecipada, e o encerramento seguirá o prazo original do calendário{obs_data}.</p>"""

            bloco_prazos_texto = f"""Próximos prazos:

{data_continuidade} – Confirmação de continuidade
{data_submissao} – Último dia para submissão da MONOGRAFIA APROVADA PELO ORIENTADOR e do TERMO DE SOLICITAÇÃO DE AVALIAÇÃO DA MONOGRAFIA

Observação importante:

Caso você envie o seu TCC para análise e ele seja aprovado pelo orientador antes da data prevista para confirmação de continuidade, a funcionalidade terá sua abertura antecipada, e o encerramento seguirá o prazo original do calendário{obs_data}."""

            rodape_html = """<p>---<br>
Portal TCC<br>
Esta é uma notificação automática. Para mais informações, acesse o sistema.</p>"""

            rodape_texto = """---
Portal TCC
Esta é uma notificação automática. Para mais informações, acesse o sistema."""

            # === 1. Email para o ALUNO ===
            corpo_html_aluno = f"""
<p>Olá, {tcc.aluno.nome_completo},</p>

<p>Sua solicitação de orientação foi aprovada com sucesso.</p>

<p><b>Tema:</b> {tcc.titulo}<br>
<b>Orientador:</b> {solicitacao.professor.nome_completo}{linha_coori_html}</p>

<p>A função de envio do TCC para análise do orientador já está habilitada em sua conta.
Acesse o sistema pelo link: <a href="http://localhost:5500">http://localhost:5500</a></p>

{bloco_prazos_html}

{rodape_html}
"""

            corpo_texto_aluno = f"""Olá, {tcc.aluno.nome_completo},

Sua solicitação de orientação foi aprovada com sucesso.

Tema: {tcc.titulo}
Orientador: {solicitacao.professor.nome_completo}{linha_coori_texto}

A função de envio do TCC para análise do orientador já está habilitada em sua conta.
Acesse o sistema pelo link: http://localhost:5500

{bloco_prazos_texto}

{rodape_texto}
"""

            criar_notificacao_com_email(
                usuario=tcc.aluno,
                tipo=TipoNotificacao.SOLICITACAO_APROVADA,
                titulo='Solicitação de Orientação Aprovada',
                mensagem=f'Sua solicitação de orientação foi aprovada com sucesso. Tema: {tcc.titulo}. Orientador: {solicitacao.professor.nome_completo}.',
                campo_preferencia='aluno_aceitar_convite_orientador',
                action_url=f'/tccs/{tcc.id}',
                tcc_id=tcc.id,
                prioridade=PrioridadeNotificacao.ALTA,
                corpo_html_customizado=corpo_html_aluno,
                corpo_texto_customizado=corpo_texto_aluno
            )

            # === 2. Email para o ORIENTADOR ===
            def _gerar_email_orientacao(nome_destinatario, tipo_orientacao):
                """Gera HTML e texto para email de orientação/co-orientação aprovada."""
                corpo_html = f"""
<p>Olá, {nome_destinatario},</p>

<p>Uma nova {tipo_orientacao} foi aprovada.</p>

<p><b>Tema:</b> {tcc.titulo}<br>
<b>Orientador:</b> {solicitacao.professor.nome_completo}{linha_coori_html}</p>

{bloco_prazos_html}

{rodape_html}
"""
                corpo_texto = f"""Olá, {nome_destinatario},

Uma nova {tipo_orientacao} foi aprovada.

Tema: {tcc.titulo}
Orientador: {solicitacao.professor.nome_completo}{linha_coori_texto}

{bloco_prazos_texto}

{rodape_texto}
"""
                return corpo_html, corpo_texto

            html_ori, texto_ori = _gerar_email_orientacao(
                solicitacao.professor.nome_completo, 'orientação'
            )
            criar_notificacao_com_email(
                usuario=solicitacao.professor,
                tipo=TipoNotificacao.SOLICITACAO_APROVADA,
                titulo='Nova Orientação Aprovada',
                mensagem=f'Uma nova orientação foi aprovada. Tema: {tcc.titulo}. Aluno: {tcc.aluno.nome_completo}.',
                campo_preferencia='prof_convite_orientacao',
                action_url=f'/tccs/{tcc.id}',
                tcc_id=tcc.id,
                prioridade=PrioridadeNotificacao.ALTA,
                corpo_html_customizado=html_ori,
                corpo_texto_customizado=texto_ori
            )

            # === 3. Email para o CO-ORIENTADOR (se existir e estiver cadastrado) ===
            if tcc.coorientador:
                html_coori, texto_coori = _gerar_email_orientacao(
                    tcc.coorientador.nome_completo, 'co-orientação'
                )
                criar_notificacao_com_email(
                    usuario=tcc.coorientador,
                    tipo=TipoNotificacao.SOLICITACAO_APROVADA,
                    titulo='Nova Co-orientação Aprovada',
                    mensagem=f'Uma nova co-orientação foi aprovada. Tema: {tcc.titulo}. Aluno: {tcc.aluno.nome_completo}.',
                    campo_preferencia=_pref(tcc.coorientador, 'prof_convite_orientacao'),
                    action_url=f'/tccs/{tcc.id}',
                    tcc_id=tcc.id,
                    prioridade=PrioridadeNotificacao.ALTA,
                    corpo_html_customizado=html_coori,
                    corpo_texto_customizado=texto_coori
                )

        return Response({
            'solicitacao': SolicitacaoOrientacaoSerializer(solicitacao, context={'request': request}).data,
            'tcc': TCCSerializer(tcc, context={'request': request}).data,
            'message': 'Solicitacao aprovada pelo coordenador com sucesso'
        })

    @action(detail=True, methods=['post'], permission_classes=[IsCoordenador])
    def recusar(self, request, pk=None):
        """
        POST /api/solicitacoes/{id}/recusar/
        Coordenador recusa solicitacao de orientacao e remove o TCC.
        Requer parecer obrigatório no body.
        Remove completamente o TCC, documentos, eventos e arquivos físicos.
        Cria registro de recusa para exibir ao aluno.
        """
        from django.db import transaction
        from .models import HistoricoRecusa

        solicitacao = self.get_object()

        if solicitacao.status != StatusSolicitacao.PENDENTE:
            return Response(
                {'detail': 'Apenas solicitacoes pendentes podem ser recusadas'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Parecer opcional
        parecer = request.data.get('parecer', '').strip()

        # Capturar dados antes de deletar
        tcc = solicitacao.tcc
        tcc_id = tcc.id
        tcc_titulo = tcc.titulo
        solicitacao_id = solicitacao.id
        aluno = tcc.aluno
        professor_nome = solicitacao.professor.nome_completo
        coordenador = request.user
        coordenador_nome = coordenador.nome_completo if hasattr(coordenador, 'nome_completo') else str(coordenador)

        # Usar transação atômica para garantir consistência
        with transaction.atomic():
            # Criar/atualizar HistoricoRecusa antes de deletar o TCC
            historico_recusa, created = HistoricoRecusa.objects.update_or_create(
                aluno=aluno,
                defaults={
                    'parecer': parecer,
                    'coordenador': coordenador,
                    'coordenador_nome': coordenador_nome,
                    'recusado_em': timezone.now()
                }
            )

            # Iterar pelos documentos e apagar os arquivos físicos
            for documento in tcc.documentos.all():
                if documento.arquivo:
                    try:
                        # Deletar arquivo físico do disco
                        documento.arquivo.delete(save=False)
                    except Exception as e:
                        # Log error but continue (file might already be deleted)
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.warning(f'Erro ao deletar arquivo do documento {documento.id}: {str(e)}')

            # Deletar TCC (cascade remove documentos, eventos, timeline e solicitações)
            tcc.delete()

            # Notificar aluno por email sobre a rejeição
            from notificacoes.services import criar_notificacao_com_email

            motivo = parecer if parecer else 'Nenhum motivo informado.'

            corpo_html = f"""
<p>Olá, {aluno.nome_completo},</p>

<p>Sua solicitação de orientação não foi aprovada.</p>

<p><b>Motivo:</b><br>
{motivo}</p>

<p><b>Tema:</b> {tcc_titulo}<br>
<b>Orientador:</b> {professor_nome}</p>

<p>Acesse o sistema em <a href="http://localhost:5500">http://localhost:5500</a> para realizar uma nova solicitação ou entre em contato com a coordenação.</p>

<p>---<br>
Portal TCC<br>
Esta é uma notificação automática.</p>
"""

            corpo_texto = f"""Olá, {aluno.nome_completo},

Sua solicitação de orientação não foi aprovada.

Motivo:
{motivo}

Tema: {tcc_titulo}
Orientador: {professor_nome}

Acesse o sistema em http://localhost:5500 para realizar uma nova solicitação ou entre em contato com a coordenação.

---
Portal TCC
Esta é uma notificação automática.
"""

            criar_notificacao_com_email(
                usuario=aluno,
                tipo=TipoNotificacao.SOLICITACAO_REJEITADA,
                titulo='Solicitação de Orientação Rejeitada',
                mensagem=f'Sua solicitação de orientação não foi aprovada. Motivo: {motivo}',
                campo_preferencia='aluno_aceitar_convite_orientador',
                action_url='/',
                prioridade=PrioridadeNotificacao.ALTA,
                corpo_html_customizado=corpo_html,
                corpo_texto_customizado=corpo_texto
            )

        return Response({
            'message': 'Solicitação recusada pelo coordenador. TCC removido.',
            'tcc_id': tcc_id,
            'solicitacao_id': solicitacao_id,
            'recusa': {
                'parecer': parecer,
                'coordenador_nome': coordenador_nome,
                'recusado_em': historico_recusa.recusado_em.isoformat()
            }
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['delete'], permission_classes=[IsAluno])
    def cancelar(self, request, pk=None):
        """
        DELETE /api/solicitacoes/{id}/cancelar/
        Aluno cancela solicitação pendente e remove completamente o TCC.
        Remove TCC, solicitação, documentos, eventos e arquivos físicos.
        """
        from django.db import transaction

        solicitacao = self.get_object()

        # Verificar se é o aluno dono
        if solicitacao.tcc.aluno != request.user:
            return Response(
                {'detail': 'Você não tem permissão para cancelar esta solicitação'},
                status=status.HTTP_403_FORBIDDEN
            )

        if solicitacao.status != StatusSolicitacao.PENDENTE:
            return Response(
                {'detail': 'Apenas solicitações pendentes podem ser canceladas'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Usar transação atômica para garantir consistência
        with transaction.atomic():
            tcc = solicitacao.tcc

            # Iterar pelos documentos e apagar os arquivos físicos
            for documento in tcc.documentos.all():
                if documento.arquivo:
                    try:
                        # Deletar arquivo físico do disco
                        documento.arquivo.delete(save=False)
                    except Exception as e:
                        # Log error but continue (file might already be deleted)
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.warning(f'Erro ao deletar arquivo do documento {documento.id}: {str(e)}')

            # Deletar TCC (cascade remove documentos, eventos, timeline e solicitações)
            tcc.delete()

        # Retornar 204 NO CONTENT (sem payload)
        return Response(status=status.HTTP_204_NO_CONTENT)


class DocumentoTCCViewSet(viewsets.ModelViewSet):
    """ViewSet para gerenciar Documentos do TCC."""

    queryset = DocumentoTCC.objects.all()
    serializer_class = DocumentoTCCSerializer
    permission_classes = [IsAuthenticated, IsDocumentoTCCRelated]

    def get_queryset(self):
        """Filtrar documentos conforme acesso ao TCC."""
        usuario = self.request.user

        if usuario.tipo_usuario == 'COORDENADOR':
            return DocumentoTCC.objects.all()

        elif usuario.tipo_usuario == 'ALUNO':
            return DocumentoTCC.objects.filter(tcc__aluno=usuario)

        elif usuario.tipo_usuario in ['PROFESSOR', 'COORDENADOR']:
            return DocumentoTCC.objects.filter(
                Q(tcc__orientador=usuario) | Q(tcc__coorientador=usuario)
            )

        return DocumentoTCC.objects.none()

    def perform_create(self, serializer):
        """Validar relação com TCC e criar evento."""
        tcc = serializer.validated_data.get('tcc')
        usuario = self.request.user

        # Validar que usuário tem permissão para enviar documento neste TCC
        tem_permissao = False

        if usuario.tipo_usuario == 'COORDENADOR':
            tem_permissao = True
        elif usuario.tipo_usuario == 'ALUNO' and tcc.aluno == usuario:
            tem_permissao = True
        elif usuario.tipo_usuario in ['PROFESSOR', 'COORDENADOR']:
            if tcc.orientador == usuario or tcc.coorientador == usuario:
                tem_permissao = True

        if not tem_permissao:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Você não tem permissão para enviar documentos neste TCC')

        # Validar prazos baseados no tipo de documento (coordenador bypassa)
        if usuario.tipo_usuario != 'COORDENADOR':
            tipo_documento = serializer.validated_data.get('tipo_documento')
            permissoes = calcular_permissoes_tcc(tcc)

            # Documentos iniciais
            if tipo_documento in [TipoDocumento.PLANO_DESENVOLVIMENTO, TipoDocumento.TERMO_ACEITE]:
                if not permissoes['pode_enviar_documentos_iniciais']:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied('Período de envio de documentos iniciais encerrado. Solicite liberação ao coordenador.')

            # Monografia
            elif tipo_documento == TipoDocumento.MONOGRAFIA:
                if not permissoes['pode_enviar_monografia']:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied('Período de envio de monografia encerrado. Solicite liberação ao coordenador.')

            # Termo de Solicitação de Avaliação
            elif tipo_documento == TipoDocumento.TERMO_SOLICITACAO_AVALIACAO:
                if not permissoes['pode_solicitar_avaliacao']:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied('Período de solicitação de avaliação encerrado. Solicite liberação ao coordenador.')

        documento = serializer.save(enviado_por=self.request.user)

        # Criar evento de upload
        # Nota: .capitalize() padroniza para sentence case (apenas primeira letra maiúscula).
        # Eventos antigos continuam com formato original - apenas novos seguem este padrão.
        tipo_display = documento.get_tipo_documento_display().capitalize()
        EventoTimeline.objects.create(
            tcc=documento.tcc,
            usuario=self.request.user,
            tipo_evento=TipoEvento.UPLOAD_DOCUMENTO,
            descricao=f'Documento enviado: {tipo_display} (v{documento.versao})',
            detalhes_json={
                'documento_id': documento.id,
                'tipo': documento.tipo_documento,
                'versao': documento.versao
            },
            visibilidade=Visibilidade.TODOS
        )

        # Criar notificações COM e-mail baseadas no tipo de documento
        from notificacoes.services import criar_notificacao_com_email, criar_notificacao_em_massa_com_email

        # Monografia enviada - notificar orientador (e co-orientador se existir)
        if documento.tipo_documento == TipoDocumento.MONOGRAFIA and usuario.tipo_usuario == 'ALUNO':
            from definicoes.models import CalendarioSemestre

            calendario = CalendarioSemestre.obter_calendario_atual()
            data_submissao = calendario.submissao_monografia_fim.strftime('%d/%m/%Y') if calendario and calendario.submissao_monografia_fim else 'Não definida'

            # Dados do co-orientador (se existir)
            tem_coorientador = bool(tcc.coorientador_nome or tcc.coorientador)
            coorientador_nome = ''
            if tcc.coorientador:
                coorientador_nome = tcc.coorientador.nome_completo
            elif tcc.coorientador_nome:
                coorientador_nome = tcc.coorientador_nome

            linha_coori_html = f'<br><b>Co-orientador:</b> {coorientador_nome}' if tem_coorientador else ''
            linha_coori_texto = f'\nCo-orientador: {coorientador_nome}' if tem_coorientador else ''

            rodape_html = """<p>---<br>
Portal TCC<br>
Esta é uma notificação automática. Para mais informações, acesse o sistema.</p>"""

            rodape_texto = """---
Portal TCC
Esta é uma notificação automática. Para mais informações, acesse o sistema."""

            def _gerar_email_monografia(nome_destinatario):
                """Gera HTML e texto para email de monografia recebida."""
                corpo_html = f"""
<p>Olá, {nome_destinatario},</p>

<p>O aluno <b>{usuario.nome_completo}</b> enviou a monografia para análise.</p>

<p><b>Tema:</b> {tcc.titulo}<br>
<b>Orientador:</b> {tcc.orientador.nome_completo}{linha_coori_html}<br>
<b>Versão:</b> {documento.versao}</p>

<p>Acesse o sistema para revisar: <a href="http://localhost:5500/tccs/{tcc.id}">http://localhost:5500/tccs/{tcc.id}</a></p>

<p><b>📅 Prazo para submissão da monografia aprovada:</b> {data_submissao}</p>

{rodape_html}
"""
                corpo_texto = f"""Olá, {nome_destinatario},

O aluno {usuario.nome_completo} enviou a monografia para análise.

Tema: {tcc.titulo}
Orientador: {tcc.orientador.nome_completo}{linha_coori_texto}
Versão: {documento.versao}

Acesse o sistema para revisar: http://localhost:5500/tccs/{tcc.id}

Prazo para submissão da monografia aprovada: {data_submissao}

{rodape_texto}
"""
                return corpo_html, corpo_texto

            # Email para o orientador
            html_ori, texto_ori = _gerar_email_monografia(tcc.orientador.nome_completo)
            criar_notificacao_com_email(
                usuario=tcc.orientador,
                tipo=TipoNotificacao.DOCUMENTO_ENVIADO,
                titulo='Nova Monografia Recebida',
                mensagem=f'Aluno {usuario.nome_completo} enviou uma nova versão da monografia (v{documento.versao}).',
                campo_preferencia='prof_receber_monografia',
                action_url=f'/tccs/{tcc.id}',
                tcc_id=tcc.id,
                prioridade=PrioridadeNotificacao.ALTA,
                corpo_html_customizado=html_ori,
                corpo_texto_customizado=texto_ori
            )

            # Email para o co-orientador (se existir e estiver cadastrado)
            if tcc.coorientador:
                html_coori, texto_coori = _gerar_email_monografia(tcc.coorientador.nome_completo)
                criar_notificacao_com_email(
                    usuario=tcc.coorientador,
                    tipo=TipoNotificacao.DOCUMENTO_ENVIADO,
                    titulo='Nova Monografia Recebida',
                    mensagem=f'Aluno {usuario.nome_completo} enviou uma nova versão da monografia (v{documento.versao}).',
                    campo_preferencia=_pref(tcc.coorientador, 'prof_receber_monografia'),
                    action_url=f'/tccs/{tcc.id}',
                    tcc_id=tcc.id,
                    prioridade=PrioridadeNotificacao.ALTA,
                    corpo_html_customizado=html_coori,
                    corpo_texto_customizado=texto_coori
                )

        # Termo de solicitação enviado - notificar coordenador
        elif documento.tipo_documento == TipoDocumento.TERMO_SOLICITACAO_AVALIACAO and usuario.tipo_usuario == 'ALUNO':
            from users.models import Usuario
            coordenadores = Usuario.objects.filter(tipo_usuario='COORDENADOR')
            criar_notificacao_em_massa_com_email(
                usuarios=list(coordenadores),
                tipo=TipoNotificacao.TERMO_ENVIADO,
                titulo='Termo de Solicitação Enviado',
                mensagem=f'Aluno {usuario.nome_completo} enviou o termo de solicitação de avaliação do TCC "{tcc.titulo}".',
                campo_preferencia='coord_termo_enviado',
                action_url=f'/tccs/{tcc.id}',
                tcc_id=tcc.id,
                prioridade=PrioridadeNotificacao.NORMAL
            )

    def perform_update(self, serializer):
        """Atualizar documento e enviar notificações quando status mudar."""
        from notificacoes.services import criar_notificacao_com_email, criar_notificacao_em_massa_com_email
        from users.models import Usuario

        documento = self.get_object()
        tcc = documento.tcc
        usuario = self.request.user

        # Armazenar status anterior
        status_anterior = documento.status

        # Salvar documento atualizado
        documento = serializer.save()

        # Verificar se o status mudou
        status_novo = documento.status

        # Se status mudou para APROVADO
        if status_anterior != StatusDocumento.APROVADO and status_novo == StatusDocumento.APROVADO:
            # Criar evento na timeline
            tipo_display = documento.get_tipo_documento_display().capitalize()
            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=usuario,
                tipo_evento=TipoEvento.DOCUMENTO_APROVADO,
                descricao=f'Documento aprovado: {tipo_display} (v{documento.versao})',
                detalhes_json={
                    'documento_id': documento.id,
                    'tipo': documento.tipo_documento,
                    'versao': documento.versao
                },
                visibilidade=Visibilidade.TODOS
            )

            # Notificar aluno sobre aprovação
            criar_notificacao_com_email(
                usuario=tcc.aluno,
                tipo=TipoNotificacao.DOCUMENTO_APROVADO,
                titulo='Documento Aprovado',
                mensagem=f'Seu documento "{tipo_display}" (v{documento.versao}) foi aprovado.',
                campo_preferencia='aluno_ajuste_monografia',  # Usar mesma preferência de ajustes
                action_url=f'/tccs/{tcc.id}',
                tcc_id=tcc.id,
                prioridade=PrioridadeNotificacao.ALTA
            )

            # Se for monografia aprovada, notificar coordenador
            if documento.tipo_documento == TipoDocumento.MONOGRAFIA:
                coordenadores = Usuario.objects.filter(tipo_usuario='COORDENADOR')
                criar_notificacao_em_massa_com_email(
                    usuarios=list(coordenadores),
                    tipo=TipoNotificacao.DOCUMENTO_APROVADO,
                    titulo='Monografia Aprovada',
                    mensagem=f'Monografia do TCC "{tcc.titulo}" (aluno: {tcc.aluno.nome_completo}) foi aprovada pelo orientador.',
                    campo_preferencia='coord_monografia_aprovada',
                    action_url=f'/tccs/{tcc.id}',
                    tcc_id=tcc.id,
                    prioridade=PrioridadeNotificacao.NORMAL
                )

        # Se status mudou para REJEITADO
        elif status_anterior != StatusDocumento.REJEITADO and status_novo == StatusDocumento.REJEITADO:
            # Criar evento na timeline
            tipo_display = documento.get_tipo_documento_display().capitalize()
            feedback_text = f'. Feedback: {documento.feedback}' if documento.feedback else ''

            EventoTimeline.objects.create(
                tcc=tcc,
                usuario=usuario,
                tipo_evento=TipoEvento.DOCUMENTO_REJEITADO,
                descricao=f'Documento rejeitado: {tipo_display} (v{documento.versao}){feedback_text}',
                detalhes_json={
                    'documento_id': documento.id,
                    'tipo': documento.tipo_documento,
                    'versao': documento.versao,
                    'feedback': documento.feedback
                },
                visibilidade=Visibilidade.TODOS
            )

            # Notificar aluno sobre rejeição/ajustes necessários
            mensagem = f'Seu documento "{tipo_display}" (v{documento.versao}) necessita de ajustes.'
            if documento.feedback:
                mensagem += f' Feedback: {documento.feedback}'

            criar_notificacao_com_email(
                usuario=tcc.aluno,
                tipo=TipoNotificacao.DOCUMENTO_REJEITADO,
                titulo='Documento Rejeitado - Ajustes Necessários',
                mensagem=mensagem,
                campo_preferencia='aluno_ajuste_monografia',
                action_url=f'/tccs/{tcc.id}',
                tcc_id=tcc.id,
                prioridade=PrioridadeNotificacao.ALTA
            )

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """
        GET /api/documentos/{id}/download/
        Retorna o arquivo com o nome original no header Content-Disposition.
        """
        from django.http import FileResponse
        from urllib.parse import quote

        documento = self.get_object()

        if not documento.arquivo:
            return Response(
                {'detail': 'Arquivo não encontrado.'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Codificar nome do arquivo para suportar caracteres especiais
        nome_arquivo = documento.nome_original or os.path.basename(documento.arquivo.name)
        nome_codificado = quote(nome_arquivo)

        response = FileResponse(
            documento.arquivo.open('rb'),
            content_type='application/pdf'
        )
        # Header para exibir no navegador (inline) com nome correto
        response['Content-Disposition'] = f"inline; filename*=UTF-8''{nome_codificado}"

        return response


class EventoTimelineViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet somente leitura para Timeline."""

    queryset = EventoTimeline.objects.all()
    serializer_class = EventoTimelineSerializer
    permission_classes = [IsAuthenticated, IsEventoTimelineVisible]

    def get_queryset(self):
        """Filtrar eventos conforme acesso ao TCC e visibilidade."""
        usuario = self.request.user
        queryset = EventoTimeline.objects.all()

        if usuario.tipo_usuario == 'COORDENADOR':
            # Coordenador vê todos os eventos
            return queryset

        elif usuario.tipo_usuario == 'ALUNO':
            # Aluno vê apenas eventos do seu TCC com visibilidade TODOS
            return queryset.filter(
                tcc__aluno=usuario,
                visibilidade=Visibilidade.TODOS
            )

        elif usuario.tipo_usuario in ['PROFESSOR', 'COORDENADOR']:
            # Professor vê eventos dos seus TCCs (TODOS + ORIENTADOR_COORDENADOR)
            from .constants import Visibilidade as Vis
            return queryset.filter(
                Q(tcc__orientador=usuario) | Q(tcc__coorientador=usuario)
            ).filter(
                Q(visibilidade=Vis.TODOS) | Q(visibilidade=Vis.ORIENTADOR_COORDENADOR)
            )

        return EventoTimeline.objects.none()

    @action(detail=False, methods=['get'])
    def por_tcc(self, request):
        """
        GET /api/timeline/por_tcc/?tcc_id=<id>
        Retorna eventos de um TCC específico.
        """
        tcc_id = request.query_params.get('tcc_id')
        if not tcc_id:
            return Response(
                {'detail': 'Parâmetro tcc_id é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )

        tcc = get_object_or_404(TCC, id=tcc_id)

        # Verificar permissão de acesso ao TCC
        usuario = request.user
        if usuario.tipo_usuario == 'ALUNO' and tcc.aluno != usuario:
            return Response(
                {'detail': 'Você não tem permissão para acessar este TCC'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Filtrar eventos
        eventos = self.get_queryset().filter(tcc=tcc).order_by('-timestamp')
        serializer = self.get_serializer(eventos, many=True)
        return Response(serializer.data)
